from openpyxl.utils import column_index_from_string, get_column_letter
from tkinter import Tk, filedialog, messagebox, BooleanVar, StringVar, ttk
from openpyxl.styles import Alignment, Font, PatternFill
from collections import defaultdict, deque
from openpyxl.cell.cell import MergedCell
from datetime import datetime, date
from openpyxl import load_workbook as _openpyxl_load_workbook
from send2trash import send2trash
from zipfile import BadZipFile, ZipFile
import tkinter.font as tkfont
from pathlib import Path
import tkinter as tk
import threading
import queue
import os
import sys


class FileAccessError(Exception):
    title = "Файл зайнятий"


_FILE_ACCESS_WINERRORS = {5, 32, 33}


def _is_file_access_error(exc: BaseException) -> bool:
    if isinstance(exc, PermissionError) or getattr(exc, "winerror", None) in _FILE_ACCESS_WINERRORS:
        return True
    text = str(exc).casefold()
    return any(
        marker in text
        for marker in (
            "permission denied",
            "access is denied",
            "used by another process",
            "file is in use",
            "locked for editing",
            "файл використовується",
            "файл зайнятий",
        )
    )


def _permission_error_path(exc: BaseException, fallback: Path | None = None) -> Path | None:
    for attr in ("filename2", "filename"):
        value = getattr(exc, attr, None)
        if value:
            try:
                return Path(value)
            except Exception:
                pass
    return Path(fallback) if fallback is not None else None


def _file_access_message(path: Path | None, action: str, exc: BaseException | None = None) -> str:
    path_text = str(path) if path else "невідомий файл"
    return (
        f"Не вдалося {action}.\n\n"
        f"Файл:\n{path_text}\n\n"
        "Причина: файл використовується іншим процесом.\n"
        "Закрийте цей файл, чи процес що його використовує..."
    )


def _raise_file_access_error(exc: BaseException, fallback: Path | None, action: str) -> None:
    path = _permission_error_path(exc, fallback)
    raise FileAccessError(_file_access_message(path, action, exc)) from exc


def ensure_file_available_for_write(path: Path, action: str = "записати файл") -> None:
    path = Path(path)
    if not path.exists():
        return
    try:
        with path.open("r+b"):
            pass
    except OSError as exc:
        if _is_file_access_error(exc):
            _raise_file_access_error(exc, path, action)
        raise


def load_workbook(filename, *args, **kwargs):
    try:
        return _openpyxl_load_workbook(filename, *args, **kwargs)
    except OSError as exc:
        if _is_file_access_error(exc):
            try:
                fallback_path = Path(filename)
            except TypeError:
                fallback_path = None
            _raise_file_access_error(exc, fallback_path, "відкрити Excel-файл")
        raise


def install_frozen_executable_icon(root, retry_ms: int = 250) -> None:
    if sys.platform != "win32" or not getattr(sys, "frozen", False):
        return

    _set_frozen_executable_icon(root)
    try:
        root.after(retry_ms, lambda: _set_frozen_executable_icon(root))
    except Exception:
        pass


def _set_frozen_executable_icon(root) -> None:
    if sys.platform != "win32" or not getattr(sys, "frozen", False):
        return

    try:
        import ctypes
        from ctypes import wintypes

        wm_seticon = 0x0080
        icon_small = 0
        icon_big = 1
        gclp_hicon = -14
        gclp_hiconsm = -34

        shell32 = ctypes.WinDLL("shell32", use_last_error=True)
        user32 = ctypes.WinDLL("user32", use_last_error=True)
        long_ptr = (
            ctypes.c_longlong if ctypes.sizeof(ctypes.c_void_p) == 8 else ctypes.c_long
        )

        shell32.ExtractIconExW.argtypes = (
            wintypes.LPCWSTR,
            ctypes.c_int,
            ctypes.POINTER(wintypes.HICON),
            ctypes.POINTER(wintypes.HICON),
            wintypes.UINT,
        )
        shell32.ExtractIconExW.restype = wintypes.UINT
        user32.SendMessageW.argtypes = (
            wintypes.HWND,
            wintypes.UINT,
            wintypes.WPARAM,
            wintypes.LPARAM,
        )
        user32.SendMessageW.restype = long_ptr
        user32.GetParent.argtypes = (wintypes.HWND,)
        user32.GetParent.restype = wintypes.HWND

        try:
            set_class_icon = user32.SetClassLongPtrW
        except AttributeError:
            set_class_icon = user32.SetClassLongW
        set_class_icon.argtypes = (wintypes.HWND, ctypes.c_int, long_ptr)
        set_class_icon.restype = long_ptr

        icon_handles = getattr(root, "_frozen_executable_icon_handles", None)
        if icon_handles is None:
            large_icons = (wintypes.HICON * 1)()
            small_icons = (wintypes.HICON * 1)()
            icon_count = shell32.ExtractIconExW(
                sys.executable,
                0,
                large_icons,
                small_icons,
                1,
            )
            if icon_count <= 0:
                return
            icon_handles = (small_icons[0], large_icons[0])
            root._frozen_executable_icon_handles = icon_handles
        small_icon, large_icon = icon_handles

        root.update_idletasks()
        hwnds = []
        for raw_hwnd in (root.winfo_id(), root.wm_frame()):
            try:
                hwnd = int(str(raw_hwnd), 0)
            except (TypeError, ValueError):
                continue
            while hwnd and hwnd not in hwnds:
                hwnds.append(hwnd)
                hwnd = user32.GetParent(hwnd)

        for hwnd in hwnds:
            if small_icon:
                user32.SendMessageW(hwnd, wm_seticon, icon_small, small_icon)
                set_class_icon(hwnd, gclp_hiconsm, small_icon)
            if large_icon:
                user32.SendMessageW(hwnd, wm_seticon, icon_big, large_icon)
                set_class_icon(hwnd, gclp_hicon, large_icon)
    except Exception:
        pass


def install_dark_title_bar(window, retry_ms: int = 80) -> None:
    if sys.platform != "win32":
        return

    def _apply() -> None:
        try:
            import ctypes

            hwnd = ctypes.windll.user32.GetParent(window.winfo_id())
            value = ctypes.c_int(1)
            for attr in (20, 19):
                ctypes.windll.dwmapi.DwmSetWindowAttribute(
                    hwnd,
                    attr,
                    ctypes.byref(value),
                    ctypes.sizeof(value),
                )
        except Exception:
            pass

    try:
        window.update_idletasks()
        _apply()
        window.after(retry_ms, _apply)
    except Exception:
        pass
import calendar
import shutil
import time
import re

# ===================== LOG + TIMERS =====================

_RUN_T0 = None

def _ts() -> str:
    return datetime.now().strftime("%H:%M:%S")

def log(msg: str):
    print(f"[{_ts()}] {msg}")

def run_start():
    global _RUN_T0
    _RUN_T0 = time.perf_counter()
    log("=== START RewardsAfter ===")

def run_end():
    if _RUN_T0 is None:
        log("=== END RewardsAfter (no timer) ===")
        return
    dt = time.perf_counter() - _RUN_T0
    log(f"=== END RewardsAfter === total={dt:.2f}s ({dt/60:.2f} min)")


def _assert_valid_xlsx_package(path: Path) -> None:
    try:
        with ZipFile(path) as zf:
            bad_member = zf.testzip()
    except BadZipFile as exc:
        raise ValueError(f"Збережений Excel-файл має пошкоджений ZIP-контейнер: {path}") from exc

    if bad_member:
        raise ValueError(f"Збережений Excel-файл має пошкоджену частину ZIP: {bad_member}")


def _safe_save_workbook(wb, target_xlsx: Path) -> None:
    target_xlsx = Path(target_xlsx)
    tmp_path = target_xlsx.with_name(f".{target_xlsx.stem}.{os.getpid()}.tmp{target_xlsx.suffix}")
    ensure_file_available_for_write(target_xlsx, "перезаписати Excel-файл")

    try:
        try:
            if tmp_path.exists():
                tmp_path.unlink()
        except OSError as exc:
            if _is_file_access_error(exc):
                _raise_file_access_error(exc, tmp_path, "прибрати старий тимчасовий Excel-файл")
            raise

        try:
            wb.save(tmp_path)
        except OSError as exc:
            if _is_file_access_error(exc):
                _raise_file_access_error(exc, tmp_path, "створити тимчасовий Excel-файл")
            raise

        _assert_valid_xlsx_package(tmp_path)
        try:
            tmp_path.replace(target_xlsx)
        except OSError as exc:
            if _is_file_access_error(exc):
                _raise_file_access_error(exc, target_xlsx, "перезаписати Excel-файл")
            raise
    except Exception:
        try:
            if tmp_path.exists():
                tmp_path.unlink()
        except Exception:
            pass
        raise


def save_target_workbook(wb, target_xlsx: Path) -> None:
    _safe_save_workbook(wb, target_xlsx)

# ===================== GUI (ONE WINDOW) =====================

def format_path_for_entry_display(path: str) -> str:
    path = path.strip()
    if not path:
        return ""

    normalized = Path(path)
    drive_prefix = ""
    if normalized.drive and len(normalized.drive) == 2 and normalized.drive[1] == ":":
        drive_prefix = f"{normalized.drive}/"

    name = normalized.name
    parent_name = normalized.parent.name
    if parent_name and name:
        return f"{drive_prefix}.../{parent_name}/{name}"
    if name:
        return f"{drive_prefix}{name}" if drive_prefix else f".../{name}"
    if drive_prefix:
        return drive_prefix
    return path.replace("\\", "/")


def fit_path_for_entry_display(path: str, entry) -> str:
    display_path = format_path_for_entry_display(path)
    if not display_path:
        return ""

    entry.update_idletasks()
    font_name = entry.cget("font") or "TkDefaultFont"
    try:
        font = tkfont.nametofont(font_name)
    except Exception:
        return display_path

    available_width = max(120, entry.winfo_width() - 14)
    if font.measure(display_path) <= available_width:
        return display_path

    parts = display_path.split("/")
    if len(parts) >= 3:
        prefix = "/".join(parts[:-1]) + "/..."
        shortened_name = parts[-1]
        while len(shortened_name) > 1 and font.measure(prefix + shortened_name) > available_width:
            shortened_name = shortened_name[1:]
        return prefix + shortened_name

    ellipsis = "..."
    shortened_path = display_path
    while len(shortened_path) > 1 and font.measure(ellipsis + shortened_path) > available_width:
        shortened_path = shortened_path[1:]
    return ellipsis + shortened_path


def refresh_path_entry_display(source_var: StringVar, display_var: StringVar, entry) -> None:
    try:
        exists = bool(entry.winfo_exists())
    except Exception:
        exists = False
    if not exists:
        return
    display_var.set(fit_path_for_entry_display(source_var.get(), entry))


DEFAULT_START_COL_LETTER = "F"
DEFAULT_START_COL_IDX = column_index_from_string(DEFAULT_START_COL_LETTER)
HEADER_PERIOD_RE = re.compile(r"\b(20\d{2})-(0[1-9]|1[0-2])\b")
HEADER_PERIOD_SCAN_SHEETS = ("100 000", "упр", "70 000", "100", "70", "30", "0")


def _parse_header_year_month(value) -> tuple[int, int] | None:
    if value is None:
        return None
    match = HEADER_PERIOD_RE.search(str(value))
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


def detect_template_period(
    target_xlsx: Path,
    preferred_sheets: tuple[str, ...] = HEADER_PERIOD_SCAN_SHEETS,
) -> tuple[int, int]:
    wb = load_workbook(target_xlsx, read_only=True, data_only=True)
    try:
        ordered_sheet_names: list[str] = []
        seen: set[str] = set()
        for sheet_name in (*preferred_sheets, *wb.sheetnames):
            if sheet_name in wb.sheetnames and sheet_name not in seen:
                ordered_sheet_names.append(sheet_name)
                seen.add(sheet_name)

        for sheet_name in ordered_sheet_names:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=5, values_only=True):
                for value in row:
                    parsed = _parse_header_year_month(value)
                    if parsed is not None:
                        return parsed
    finally:
        wb.close()

    raise ValueError("Не знайшов формат YYYY-MM у заголовку шаблону (A1:E3).")


def detect_template_context(
    target_xlsx: Path,
    preferred_sheets: tuple[str, ...] = HEADER_PERIOD_SCAN_SHEETS,
) -> tuple[int, int, int]:
    year, month = detect_template_period(target_xlsx, preferred_sheets=preferred_sheets)
    return DEFAULT_START_COL_IDX, year, month


def configure_launch_styles(root):
    colors = {
        "window": "#F4F7FB",
        "panel": "#FFFFFF",
        "border": "#D6DEE8",
        "text": "#15202B",
        "muted": "#617084",
        "accent": "#0F766E",
        "accent_dark": "#0B5D56",
        "header": "#103C3A",
        "header_badge": "#0F766E",
    }

    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except Exception:
        pass

    try:
        style.configure("LaunchRoot.TFrame", background=colors["window"])
        style.configure("LaunchPanel.TFrame", background=colors["panel"])
        style.configure("LaunchSection.TLabel", background=colors["panel"], foreground=colors["text"], font=("Segoe UI", 12, "bold"))
        style.configure("LaunchField.TLabel", background=colors["panel"], foreground=colors["text"], font=("Segoe UI", 10, "bold"))
        style.configure("LaunchMuted.TLabel", background=colors["panel"], foreground=colors["muted"], font=("Segoe UI", 9))
        style.configure("ProgressHeader.TLabel", background=colors["panel"], foreground=colors["text"], font=("Segoe UI", 12, "bold"))
        style.configure("ProgressBody.TLabel", background=colors["panel"], foreground=colors["text"], font=("Segoe UI", 10))
        style.configure("ProgressFile.TLabel", background=colors["panel"], foreground=colors["muted"], font=("Segoe UI", 9))
        style.configure("ProgressCount.TLabel", background=colors["panel"], foreground=colors["accent_dark"], font=("Segoe UI", 10, "bold"))
        style.configure(
            "Launch.Horizontal.TProgressbar",
            background=colors["header_badge"],
            troughcolor="#E8EEF5",
            bordercolor=colors["border"],
            lightcolor=colors["header_badge"],
            darkcolor=colors["header_badge"],
            thickness=12,
        )
        style.configure(
            "Launch.TEntry",
            fieldbackground="#FFFFFF",
            foreground=colors["text"],
            bordercolor=colors["border"],
            lightcolor=colors["border"],
            darkcolor=colors["border"],
            padding=(8, 6),
        )
        style.map(
            "Launch.TEntry",
            fieldbackground=[("readonly", "#FFFFFF")],
            foreground=[("readonly", colors["text"])],
        )
        style.configure(
            "LaunchBrowse.TButton",
            font=("Segoe UI", 10, "bold"),
            padding=(18, 10),
            foreground="#FFFFFF",
            background="#053D39",
            bordercolor=colors["accent_dark"],
            lightcolor=colors["accent_dark"],
            darkcolor=colors["accent_dark"],
            focuscolor=colors["accent_dark"],
        )
        style.map(
            "LaunchBrowse.TButton",
            background=[("pressed", colors["accent_dark"]), ("active", colors["accent_dark"])],
            foreground=[("pressed", "#FFFFFF"), ("active", "#FFFFFF")],
        )
        style.configure(
            "LaunchPrimary.TButton",
            font=("Segoe UI", 10, "bold"),
            padding=(18, 10),
            foreground="#FFFFFF",
            background="#053D39",
            bordercolor=colors["accent_dark"],
            lightcolor=colors["accent_dark"],
            darkcolor=colors["accent_dark"],
            focuscolor=colors["accent_dark"],
        )
        style.map(
            "LaunchPrimary.TButton",
            background=[("pressed", colors["accent_dark"]), ("active", colors["accent_dark"])],
            foreground=[("pressed", "#FFFFFF"), ("active", "#FFFFFF")],
        )
        style.configure(
            "LaunchHero.TButton",
            font=("Segoe UI", 11, "bold"),
            padding=(20, 12),
            foreground="#FFFFFF",
            background="#053D39",
            bordercolor=colors["accent_dark"],
            lightcolor=colors["accent_dark"],
            darkcolor=colors["accent_dark"],
            focuscolor=colors["accent_dark"],
        )
        style.map(
            "LaunchHero.TButton",
            background=[("pressed", colors["accent_dark"]), ("active", colors["accent_dark"])],
            foreground=[("pressed", "#FFFFFF"), ("active", "#FFFFFF")],
        )
    except Exception:
        pass

    return colors


def _widget_exists(widget) -> bool:
    try:
        return bool(widget.winfo_exists())
    except Exception:
        return False


class ProgressWindow:
    def __init__(self, owner: Tk, subtitle: str = "Заповнення листів 100 / 70 / 30 / 0 виконується..."):
        self.close_requested = False
        self.dialog = tk.Toplevel(owner)
        self.dialog.withdraw()
        install_frozen_executable_icon(self.dialog)
        self.dialog.title("RewardsAfter - Обробка")
        self.dialog.resizable(False, False)
        self.dialog.attributes("-topmost", True)
        self.dialog.protocol("WM_DELETE_WINDOW", self.request_close)
        colors = configure_launch_styles(self.dialog)
        self.dialog.configure(bg=colors["window"])
        install_dark_title_bar(self.dialog)

        container = ttk.Frame(self.dialog, style="LaunchRoot.TFrame")
        container.pack(fill="both", expand=True)

        header = tk.Frame(container, bg=colors["header"], padx=22, pady=10)
        header.pack(fill="x")
        header.grid_columnconfigure(0, weight=1)
        tk.Label(
            header,
            text="RewardsAfter",
            bg=colors["header"],
            fg="#FFFFFF",
            font=("Segoe UI Semibold", 15),
            anchor="w",
        ).grid(row=0, column=0, sticky="w")
        tk.Label(
            header,
            text=subtitle,
            bg=colors["header"],
            fg="#D7FBF5",
            font=("Segoe UI", 9),
            anchor="w",
        ).grid(row=1, column=0, sticky="w", pady=(3, 4))
        badge_shell = tk.Frame(header, bg=colors["accent_dark"], padx=1, pady=1)
        badge = tk.Label(
            badge_shell,
            text="ОБРОБКА",
            bg="#053D39",
            fg="#FFFFFF",
            font=("Segoe UI", 12, "bold"),
            padx=10,
            pady=5,
        )
        badge.pack()
        badge_shell.place(relx=1.0, x=5, y=5, anchor="ne")

        body = ttk.Frame(container, style="LaunchRoot.TFrame", padding=(18, 16, 18, 18))
        body.pack(fill="both", expand=True)
        panel_shell = tk.Frame(body, bg=colors["border"], padx=1, pady=1)
        panel_shell.pack(fill="both", expand=True)
        frame = ttk.Frame(panel_shell, style="LaunchPanel.TFrame", padding=(16, 14, 16, 12))
        frame.pack(fill="both", expand=True)
        frame.grid_columnconfigure(0, weight=1)

        self.header_var = tk.StringVar(value="Підготовка...")
        self.detail_var = tk.StringVar(value="Будь ласка, зачекайте")
        self.file_var = tk.StringVar(value="")
        self.count_var = tk.StringVar(value="0 / 0")

        ttk.Label(frame, textvariable=self.header_var, style="ProgressHeader.TLabel").grid(
            row=0, column=0, padx=(0, 12), sticky="w"
        )
        ttk.Label(frame, textvariable=self.count_var, style="ProgressCount.TLabel").grid(
            row=0, column=1, sticky="e"
        )
        ttk.Label(frame, textvariable=self.detail_var, style="ProgressBody.TLabel", wraplength=500).grid(
            row=1, column=0, columnspan=2, pady=(8, 6), sticky="w"
        )
        self.progress = ttk.Progressbar(
            frame,
            orient="horizontal",
            mode="determinate",
            maximum=100,
            style="Launch.Horizontal.TProgressbar",
        )
        self.progress.grid(row=2, column=0, columnspan=2, pady=(6, 0), sticky="we")
        self.dialog.update_idletasks()
        width = max(560, self.dialog.winfo_reqwidth())
        height = max(222, self.dialog.winfo_reqheight())
        self.dialog.geometry(f"{width}x{height}")
        self.dialog.deiconify()
        self.dialog.lift()
        self.refresh()

    def update(self, header=None, detail=None, current=None, total=None, file_name=None):
        try:
            if self.close_requested or not _widget_exists(self.dialog):
                return
            if header is not None:
                self.header_var.set(header)
            if detail is not None:
                self.detail_var.set(detail)
            if file_name is not None:
                self.file_var.set(file_name)
            if current is not None and total is not None:
                safe_total = total if total > 0 else 1
                pct = max(0.0, min(100.0, (current / safe_total) * 100.0))
                self.progress["value"] = pct
                self.count_var.set(f"{current} / {total}")
            self.refresh()
        except tk.TclError:
            return

    def refresh(self):
        try:
            if self.close_requested or not _widget_exists(self.dialog):
                return
            self.dialog.update_idletasks()
            self.dialog.update()
        except Exception:
            pass

    def close(self):
        self.close_requested = True
        try:
            if _widget_exists(self.dialog):
                self.dialog.destroy()
        except Exception:
            pass

    def request_close(self):
        self.close()

    def show_success_then_close(self, detail: str, file_name: str = "", delay_ms: int = 1300):
        if self.close_requested or not _widget_exists(self.dialog):
            return
        self.update(
            header="Готово",
            detail=detail,
            current=1,
            total=1,
            file_name=file_name,
        )
        deadline = time.monotonic() + delay_ms / 1000.0
        while time.monotonic() < deadline:
            if not _widget_exists(self.dialog):
                return
            self.refresh()
            time.sleep(0.05)
        self.close()


class ProgressUpdateProxy:
    def __init__(self, updates: "queue.Queue[dict[str, object]]"):
        self.updates = updates

    def update(self, header=None, detail=None, current=None, total=None, file_name=None):
        self.updates.put(
            {
                "header": header,
                "detail": detail,
                "current": current,
                "total": total,
                "file_name": file_name,
            }
        )


MIN_PROGRESS_STAGE_SECONDS = 2.0
PROGRESS_POLL_SECONDS = 0.03


class ProgressDisplayGate:
    def __init__(self, progress: ProgressWindow, min_stage_seconds: float = MIN_PROGRESS_STAGE_SECONDS):
        self.progress = progress
        self.min_stage_seconds = min_stage_seconds
        self.pending: deque[dict[str, object]] = deque()
        self.active_header = None
        self.active_since = 0.0

    def drain_updates(self, updates: "queue.Queue[dict[str, object]]") -> None:
        while True:
            try:
                self.pending.append(updates.get_nowait())
            except queue.Empty:
                return

    def has_pending(self) -> bool:
        return bool(self.pending)

    def apply_ready(self) -> bool:
        if not self.pending:
            return False

        payload = self.pending[0]
        next_header = payload.get("header") or self.active_header
        now = time.monotonic()

        if (
            self.active_header is not None
            and next_header != self.active_header
            and now - self.active_since < self.min_stage_seconds
        ):
            return False

        payload = self.pending.popleft()
        while self.pending:
            following_payload = self.pending[0]
            following_header = following_payload.get("header") or next_header
            if following_header != next_header:
                break
            payload = self.pending.popleft()

        if next_header != self.active_header:
            self.active_header = next_header
            self.active_since = time.monotonic()
        self.progress.update(**payload)
        return True


class ToggleSwitch(ttk.Frame):
    def __init__(self, owner, variable):
        super().__init__(owner)
        self.variable = variable
        self.state_var = tk.StringVar()

        style = ttk.Style(self)
        try:
            parent_style = owner.cget("style")
            if parent_style:
                self.configure(style=parent_style)
                canvas_bg = style.lookup(parent_style, "background")
            else:
                canvas_bg = ""
        except Exception:
            canvas_bg = ""
        canvas_bg = canvas_bg or style.lookup("LaunchPanel.TFrame", "background") or self.winfo_toplevel().cget("bg")
        state_label_style = "ToggleState.TLabel"
        try:
            style.configure(
                state_label_style,
                background="#F4F4F8",
                foreground="#15202B",
                font=("Segoe UI", 9),
            )
        except Exception:
            state_label_style = ""

        self.canvas = tk.Canvas(
            self,
            width=52,
            height=28,
            bd=0,
            highlightthickness=0,
            bg=canvas_bg,
            cursor="hand2",
        )
        self.canvas.grid(row=0, column=0, padx=(0, 8), sticky="w")

        self.state_label = ttk.Label(self, textvariable=self.state_var, style=state_label_style)
        self.state_label.grid(row=0, column=1, sticky="w")

        for widget in (self, self.canvas, self.state_label):
            widget.bind("<Button-1>", self.toggle)

        self.variable.trace_add("write", self._redraw)
        self._redraw()

    def toggle(self, _event=None) -> None:
        self.variable.set(not self.variable.get())

    def _redraw(self, *_args) -> None:
        is_enabled = bool(self.variable.get())
        track_color = "#0F766E" if is_enabled else "#CBD5E1"
        knob_color = "#FFFFFF"
        self.state_var.set("Увімкнено" if is_enabled else "Вимкнено")

        self.canvas.delete("all")

        left = 2
        top = 2
        right = 50
        bottom = 26
        radius = (bottom - top) // 2

        self.canvas.create_oval(left, top, left + 2 * radius, bottom, fill=track_color, outline=track_color)
        self.canvas.create_oval(right - 2 * radius, top, right, bottom, fill=track_color, outline=track_color)
        self.canvas.create_rectangle(left + radius, top, right - radius, bottom, fill=track_color, outline=track_color)

        knob_diameter = (2 * radius) - 2
        knob_left = right - knob_diameter - 1 if is_enabled else left + 1
        self.canvas.create_oval(
            knob_left,
            top + 1,
            knob_left + knob_diameter,
            bottom - 1,
            fill=knob_color,
            outline=knob_color,
        )


def run_rewards_after_with_progress(
    target_xlsx: Path,
    start_col_idx: int,
    year: int,
    month: int,
    progress: ProgressWindow,
) -> dict[str, dict[str, object]] | None:
    updates: "queue.Queue[dict[str, object]]" = queue.Queue()
    result = {"error": None, "stats": None}
    progress_proxy = ProgressUpdateProxy(updates)
    display_gate = ProgressDisplayGate(progress)

    def worker() -> None:
        try:
            backup_excel_to_trash(target_xlsx, progress=progress_proxy)
            result["stats"] = run_rewards_after(
                target_xlsx=target_xlsx,
                start_col_idx=start_col_idx,
                year=year,
                month=month,
                progress=progress_proxy,
            )
        except BaseException as exc:
            result["error"] = exc

    thread = threading.Thread(target=worker, daemon=True)
    thread.start()

    while thread.is_alive():
        display_gate.drain_updates(updates)
        display_gate.apply_ready()
        progress.refresh()
        time.sleep(PROGRESS_POLL_SECONDS)

    thread.join()
    display_gate.drain_updates(updates)
    while display_gate.has_pending():
        display_gate.apply_ready()
        progress.refresh()
        time.sleep(PROGRESS_POLL_SECONDS)

    if result["error"] is not None:
        if isinstance(result["error"], SystemExit):
            raise RuntimeError(str(result["error"]))
        raise result["error"]

    return result["stats"]


def run_rewards_money_with_progress(
    target_xlsx: Path,
    start_col_idx: int,
    year: int,
    month: int,
    progress: ProgressWindow,
) -> dict[str, dict[str, object]]:
    updates: "queue.Queue[dict[str, object]]" = queue.Queue()
    result = {"error": None, "stats": None}
    progress_proxy = ProgressUpdateProxy(updates)
    display_gate = ProgressDisplayGate(progress)

    def worker() -> None:
        try:
            backup_excel_to_trash(target_xlsx, progress=progress_proxy, total_stages=4)
            result["stats"] = run_rewards_money(
                target_xlsx=target_xlsx,
                start_col_idx=start_col_idx,
                year=year,
                month=month,
                progress=progress_proxy,
            )
        except BaseException as exc:
            result["error"] = exc

    thread = threading.Thread(target=worker, daemon=True)
    thread.start()

    while thread.is_alive():
        display_gate.drain_updates(updates)
        display_gate.apply_ready()
        progress.refresh()
        time.sleep(PROGRESS_POLL_SECONDS)

    thread.join()
    display_gate.drain_updates(updates)
    while display_gate.has_pending():
        display_gate.apply_ready()
        progress.refresh()
        time.sleep(PROGRESS_POLL_SECONDS)

    if result["error"] is not None:
        if isinstance(result["error"], SystemExit):
            raise RuntimeError(str(result["error"]))
        raise result["error"]

    return result["stats"] or {}


def ask_all_inputs_after() -> tuple[Path, int, int, int, bool]:
    root = Tk()
    root.withdraw()
    install_frozen_executable_icon(root)
    root.title("НАЛАШТУВАННЯ: RewardsAfter")
    root.attributes("-topmost", True)
    root.resizable(False, False)
    colors = configure_launch_styles(root)
    root.configure(bg=colors["window"])
    install_dark_title_bar(root)

    file_var = StringVar(value="")
    file_display_var = StringVar(value="")
    money_mode_var = BooleanVar(value=True)
    entry_width = 40

    def browse_file():
        root.lift()
        root.update()
        p = filedialog.askopenfilename(
            title="Виберіть ЦІЛЬ",
            parent=root,
            filetypes=[("Excel files", "*.xlsx;*.xlsm"), ("All files", "*.*")]
        )
        if p:
            file_var.set(p)
            refresh_path_entry_display(file_var, file_display_var, file_entry)
        root.lift()
        root.update()

    surface = ttk.Frame(root, style="LaunchRoot.TFrame")
    surface.grid(row=0, column=0, sticky="nsew")
    surface.grid_columnconfigure(0, weight=1)

    header = tk.Frame(surface, bg=colors["header"], padx=22, pady=12)
    header.grid(row=0, column=0, sticky="we")
    header.grid_columnconfigure(0, weight=1)
    tk.Label(
        header,
        text="RewardsAfter",
        bg=colors["header"],
        fg="#FFFFFF",
        font=("Segoe UI Semibold", 16),
        anchor="w",
    ).grid(row=0, column=0, sticky="w")
    tk.Label(
        header,
        text="Підготовка цільового Excel. Період визначиться автоматично.",
        bg=colors["header"],
        fg="#D7FBF5",
        font=("Segoe UI", 9),
        anchor="w",
    ).grid(row=1, column=0, sticky="w", pady=(4, 6))
    badge_shell = tk.Frame(header, bg=colors["accent_dark"], padx=1, pady=1)
    badge = tk.Label(
        badge_shell,
        text="НАЛАШТУВАННЯ",
        bg="#053D39",
        fg="#FFFFFF",
        font=("Segoe UI", 12, "bold"),
        padx=10,
        pady=5,
    )
    badge.pack()
    badge_shell.place(relx=1.0, x=5, y=5, anchor="ne")

    body = ttk.Frame(surface, style="LaunchRoot.TFrame", padding=(18, 16, 18, 18))
    body.grid(row=1, column=0, sticky="nsew")
    panel_shell = tk.Frame(body, bg=colors["border"], padx=1, pady=1)
    panel_shell.grid(row=0, column=0, sticky="we")
    container = ttk.Frame(panel_shell, style="LaunchPanel.TFrame", padding=(16, 14, 16, 12))
    container.pack(fill="both", expand=True)
    container.grid_columnconfigure(1, weight=1)

    ttk.Label(container, text="ФАЙЛ:", style="LaunchSection.TLabel").grid(
        row=0, column=0, columnspan=3, sticky="w"
    )

    ttk.Label(container, text="ЦІЛЬ (.xlsx):", style="LaunchField.TLabel").grid(row=1, column=0, sticky="e", padx=8, pady=6)
    file_entry = ttk.Entry(container, textvariable=file_display_var, width=entry_width, state="readonly", style="Launch.TEntry")
    file_entry.grid(row=1, column=1, sticky="we", padx=8, pady=6)
    ttk.Button(container, text="ОБРАТИ", width=12, command=browse_file, style="LaunchBrowse.TButton").grid(
        row=1, column=2, sticky="w", padx=8, pady=6
    )
    file_var.trace_add("write", lambda *_args: refresh_path_entry_display(file_var, file_display_var, file_entry))
    file_entry.bind("<Configure>", lambda _event: refresh_path_entry_display(file_var, file_display_var, file_entry))

    footer_line = tk.Frame(container, bg=colors["border"], height=1)
    footer_line.grid(row=2, column=0, columnspan=3, sticky="we", padx=8, pady=(12, 12))

    money_row = ttk.Frame(container, style="LaunchPanel.TFrame")
    money_row.grid(row=3, column=0, columnspan=3, sticky="w", padx=8)
    ttk.Label(money_row, text="Перерахунок грошей у AO", style="LaunchField.TLabel").grid(
        row=0, column=0, sticky="w"
    )
    ToggleSwitch(money_row, money_mode_var).grid(row=0, column=1, padx=(10, 0), sticky="w")

    ttk.Label(
        container,
        text=(
            "Коли вимкнено, виконується кінцева операція заповнення листів 100 / 70 / 30 / 0."
        ),
        style="LaunchMuted.TLabel",
        wraplength=520,
        justify="left",
    ).grid(row=4, column=0, columnspan=3, padx=8, pady=(10, 14), sticky="w")

    footer_line2 = tk.Frame(container, bg=colors["border"], height=1)
    footer_line2.grid(row=5, column=0, columnspan=3, sticky="we", padx=8, pady=(0, 14))

    result = {"ok": False, "data": None}

    def validate_and_close():
        try:
            f = file_var.get().strip()
            if not f:
                raise ValueError("Не вибрано файл-ціль.")
            target = Path(f)
            if not target.exists():
                raise ValueError("Файл не існує.")

            start_col_idx, year, month = detect_template_context(target)

            result["data"] = (target, start_col_idx, year, month, bool(money_mode_var.get()))
            result["ok"] = True
            root.destroy()

        except Exception as ex:
            messagebox.showerror("Помилка", str(ex), parent=root)

    ttk.Button(
        container,
        text="ЗАПУСТИТИ ОБРОБКУ",
        command=validate_and_close,
        style="LaunchHero.TButton",
    ).grid(
        row=6, column=0, columnspan=3, sticky="we", padx=8, pady=(0, 4)
    )

    root.update_idletasks()
    root.geometry(f"{root.winfo_reqwidth()}x{root.winfo_reqheight()}")
    root.deiconify()
    root.lift()
    root.grab_set()
    root.focus_force()
    root.mainloop()

    if not result["ok"]:
        raise SystemExit("Скасовано користувачем.")

    return result["data"]


def backup_excel_to_trash(xlsx_path: Path, progress=None, total_stages: int = 5):
    timestamp = datetime.now().strftime("%Y.%m.%d %H.%M")
    backup_name = f"{xlsx_path.stem} - {timestamp}{xlsx_path.suffix}"
    backup_path = xlsx_path.with_name(backup_name)

    try:
        if progress:
            progress.update(
                header=f"Етап 1/{total_stages}: Резервна копія",
                detail="Створюю backup цільового Excel...",
                current=0,
                total=1,
                file_name=xlsx_path.name,
            )
        try:
            shutil.copy2(xlsx_path, backup_path)
        except OSError as exc:
            if _is_file_access_error(exc):
                _raise_file_access_error(exc, xlsx_path, "зробити backup Excel-файлу")
            raise
        send2trash(str(backup_path))
        if progress:
            progress.update(
                header=f"Етап 1/{total_stages}: Резервна копія",
                detail="Backup створено та переміщено в кошик",
                current=1,
                total=1,
                file_name=backup_path.name,
            )
    except FileAccessError:
        raise
    except Exception as e:
        if progress:
            raise RuntimeError(f"Не вдалося зробити backup Excel у кошик:\n{backup_path}\n\n{e}") from e
        warning_root = Tk()
        warning_root.withdraw()
        install_frozen_executable_icon(warning_root)
        install_dark_title_bar(warning_root)
        warning_root.attributes("-topmost", True)
        try:
            messagebox.showwarning(
                "Увага",
                f"Не вдалося зробити backup Excel у кошик:\n{backup_path}\n\n{e}",
                parent=warning_root,
            )
        finally:
            warning_root.destroy()

# ===================== UTILS =====================

NO_FILL = PatternFill()
OUTPUT_START_ROW = 7
OUTPUT_SHEET_COLUMNS = {
    "100": (2, 3, 4, 5, 6),     # B..F
    "70": (2, 3, 4, 5, 6),      # B..F
    "30": (2, 3, 4, 5, 6),      # B..F
    "0": (2, 3, 4, 5, 6, 8),    # B..F, H
}
OUTPUT_CENTER_WRAP_ALIGNMENT = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True,
)
OUTPUT_DEFAULT_COLUMN_WIDTH = 8.43
OUTPUT_MIN_ROW_HEIGHT = 15.0
OUTPUT_ROW_HEIGHT_PER_LINE = 15.0
OUTPUT_ROW_HEIGHT_PADDING = 2.0
OUTPUT_MAX_ROW_HEIGHT = 180.0

def _fill_rgb(cell) -> str:
    try:
        fill = cell.fill
        if not fill or fill.patternType is None:
            return ""
        fg = getattr(fill, "fgColor", None)
        rgb = getattr(fg, "rgb", None) if fg else None
        return str(rgb).upper() if rgb else ""
    except Exception:
        return ""

def _day_to_col(start_col_idx: int, day: int) -> int:
    return start_col_idx + (day - 1)

def _compress_days_to_periods(days: list[int], year: int, month: int) -> list[tuple[date, date]]:
    if not days:
        return []
    days = sorted(set(days))
    out: list[tuple[date, date]] = []

    s = days[0]
    prev = days[0]
    for d in days[1:]:
        if d == prev + 1:
            prev = d
            continue
        out.append((date(year, month, s), date(year, month, prev)))
        s = prev = d
    out.append((date(year, month, s), date(year, month, prev)))
    return out

def _fmt_date_ua(d: date) -> str:
    return d.strftime("%d.%m.%Y")

def _find_marker_cutoff_row(ws, start_row: int, marker_col_idx: int, marker: str) -> int | None:
    m = (marker or "").casefold()
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=marker_col_idx).value
        if v and m in str(v).strip().casefold():
            return r
    return None

def _norm_pib(v) -> str:
    s = str(v or "").strip()
    if not s:
        return ""
    return " ".join(s.casefold().split())

def _build_pib_row_queues(ws, start_row: int, last_row: int, pib_col_idx: int = 5) -> dict[str, deque[int]]:
    out: dict[str, deque[int]] = defaultdict(deque)
    for r in range(start_row, last_row + 1):
        pib = _norm_pib(ws.cell(row=r, column=pib_col_idx).value)
        if pib:
            out[pib].append(r)
    return out

# ===================== COLORS =====================

OLIVE_SUFFIX = "C4D79B"   # "FFC4D79B"
GREEN_SUFFIX = "00B050"   # "FF00B050"
GREEN_SUFFIXES = (GREEN_SUFFIX, "006100")

MONEY_REQUIRED_SHEET = "100 000"
MONEY_OPTIONAL_SHEET = "упр"
MONEY_TARGET_SHEETS = (MONEY_REQUIRED_SHEET, MONEY_OPTIONAL_SHEET)
MONEY_PIB_COL_IDX = column_index_from_string("E")
MONEY_COL_LETTER = "AO"
MONEY_COL_IDX = column_index_from_string(MONEY_COL_LETTER)
MONEY_HEADER_ROWS_TO_SKIP = 5
MONEY_DEFAULT_DATA_START_ROW = MONEY_HEADER_ROWS_TO_SKIP + 1
OLIVE_MONTHLY_RATE = 30000.0
GREEN_MONTHLY_RATE = 100000.0
ACCOUNTING_UAH_FORMAT = '_-* #,##0.00\\ "грн"_-;-* #,##0.00\\ "грн"_-;_-* "-"??\\ "грн"_-;_-@_-'
MONEY_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=False)
MONEY_FONT = Font(name="Times New Roman", size=12)


def _is_money_olive_fill(cell) -> bool:
    return _fill_rgb(cell).endswith(OLIVE_SUFFIX)


def _is_money_green_fill(cell) -> bool:
    rgb = _fill_rgb(cell)
    return any(rgb.endswith(suffix) for suffix in GREEN_SUFFIXES)


def _is_person_pib(value) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    return text.casefold() != "піб"


def find_money_last_text_row(ws, col_idx: int, start_row: int) -> int:
    last = start_row - 1
    for row in range(start_row, ws.max_row + 1):
        value = ws.cell(row=row, column=col_idx).value
        if value is not None and str(value).strip() != "":
            last = row
    return last


def find_money_pib_header_row(ws) -> int | None:
    for row in range(1, min(ws.max_row, 12) + 1):
        value = ws.cell(row=row, column=MONEY_PIB_COL_IDX).value
        if str(value or "").strip().casefold() == "піб":
            return row
    return None


def detect_money_data_start_row(ws) -> int:
    header_row = find_money_pib_header_row(ws)
    if header_row is None:
        return MONEY_DEFAULT_DATA_START_ROW

    search_to = max(MONEY_DEFAULT_DATA_START_ROW, header_row + 1)
    for row in range(header_row + 1, min(ws.max_row, search_to) + 1):
        if _is_person_pib(ws.cell(row=row, column=MONEY_PIB_COL_IDX).value):
            return row

    return MONEY_DEFAULT_DATA_START_ROW


def clear_money_column_text_only(ws, start_row: int) -> int:
    cleared = 0

    for row in range(start_row, max(ws.max_row, start_row) + 1):
        # Do not use ws.cell()/iter_rows here: they create blank cells and can disturb template formatting on save.
        cell = ws._cells.get((row, MONEY_COL_IDX))
        if cell is None:
            continue
        if isinstance(cell, MergedCell):
            continue
        if cell.value not in (None, ""):
            cleared += 1
            cell.value = None

    return cleared


def calculate_row_salary(ws, row: int, start_col_idx: int, last_day: int) -> tuple[float, int, int]:
    olive_days = 0
    green_days = 0

    for day in range(1, last_day + 1):
        cell = ws.cell(row=row, column=_day_to_col(start_col_idx, day))
        if _is_money_olive_fill(cell):
            olive_days += 1
        elif _is_money_green_fill(cell):
            green_days += 1

    olive_daily = OLIVE_MONTHLY_RATE / last_day
    green_daily = GREEN_MONTHLY_RATE / last_day
    amount = round((olive_days * olive_daily) + (green_days * green_daily), 2)
    return amount, olive_days, green_days


def process_money_sheet(ws, start_col_idx: int, last_day: int, progress=None, sheet_no: int = 1, total_sheets: int = 1) -> dict[str, object]:
    data_start_row = detect_money_data_start_row(ws)
    cleared_values = clear_money_column_text_only(ws, start_row=data_start_row)
    last_row = find_money_last_text_row(ws, MONEY_PIB_COL_IDX, data_start_row)

    stats: dict[str, object] = {
        "sheet": ws.title,
        "data_start_row": data_start_row,
        "last_row": last_row,
        "cleared_values": cleared_values,
        "processed_rows": 0,
        "written_rows": 0,
        "olive_days": 0,
        "green_days": 0,
        "total_amount": 0.0,
    }

    if last_row < data_start_row:
        return stats

    total_rows = last_row - data_start_row + 1
    for idx, row in enumerate(range(data_start_row, last_row + 1), start=1):
        if progress and (idx == 1 or idx % 25 == 0 or idx == total_rows):
            progress.update(
                header=f"Етап 3/4: Аркуш {sheet_no}/{total_sheets}",
                detail=f"Рахую зарплати на аркуші '{ws.title}'...",
                current=idx,
                total=total_rows,
                file_name=f"Рядок {row}",
            )

        pib_value = ws.cell(row=row, column=MONEY_PIB_COL_IDX).value
        if not _is_person_pib(pib_value):
            continue

        amount, olive_days, green_days = calculate_row_salary(ws, row, start_col_idx, last_day)
        stats["processed_rows"] = int(stats["processed_rows"]) + 1
        if amount == 0:
            continue

        target_cell = ws.cell(row=row, column=MONEY_COL_IDX)
        if isinstance(target_cell, MergedCell):
            raise ValueError(f"Клітинка {ws.title}!{target_cell.coordinate} об'єднана, не можу записати суму.")

        target_cell.value = amount
        target_cell.number_format = ACCOUNTING_UAH_FORMAT
        target_cell.alignment = MONEY_ALIGNMENT
        target_cell.font = MONEY_FONT

        stats["written_rows"] = int(stats["written_rows"]) + 1
        stats["olive_days"] = int(stats["olive_days"]) + olive_days
        stats["green_days"] = int(stats["green_days"]) + green_days
        stats["total_amount"] = float(stats["total_amount"]) + amount

    stats["total_amount"] = round(float(stats["total_amount"]), 2)
    return stats


def run_rewards_money(target_xlsx: Path, start_col_idx: int, year: int, month: int, progress=None) -> dict[str, dict[str, object]]:
    last_day = calendar.monthrange(year, month)[1]

    if progress:
        progress.update(
            header="Етап 2/4: Завантаження Excel",
            detail="Відкриваю цільовий файл...",
            current=0,
            total=1,
            file_name=target_xlsx.name,
        )

    wb = load_workbook(target_xlsx)
    try:
        if MONEY_REQUIRED_SHEET not in wb.sheetnames:
            raise ValueError(f"Нема аркуша '{MONEY_REQUIRED_SHEET}' у файлі: {target_xlsx.name}")

        sheet_names = [sheet_name for sheet_name in MONEY_TARGET_SHEETS if sheet_name in wb.sheetnames]
        if progress:
            progress.update(
                header="Етап 2/4: Завантаження Excel",
                detail=f"Період: {year}-{month:02d}, днів: {last_day}",
                current=1,
                total=1,
                file_name=", ".join(sheet_names),
            )

        stats: dict[str, dict[str, object]] = {}
        for sheet_no, sheet_name in enumerate(sheet_names, start=1):
            stats[sheet_name] = process_money_sheet(
                wb[sheet_name],
                start_col_idx=start_col_idx,
                last_day=last_day,
                progress=progress,
                sheet_no=sheet_no,
                total_sheets=len(sheet_names),
            )

        if progress:
            progress.update(
                header="Етап 4/4: Збереження",
                detail="Зберігаю Excel-файл...",
                current=0,
                total=1,
                file_name=target_xlsx.name,
            )

        save_target_workbook(wb, target_xlsx)

        if progress:
            progress.update(
                header="Етап 4/4: Збереження",
                detail="Файл збережено",
                current=1,
                total=1,
                file_name=target_xlsx.name,
            )

        return stats
    finally:
        wb.close()


def format_money_success_message(stats: dict[str, dict[str, object]]) -> str:
    if not stats:
        return "RewardsAfter: перерахунок грошей завершено."

    parts = []
    for sheet_name, sheet_stats in stats.items():
        written = int(sheet_stats.get("written_rows", 0))
        amount = float(sheet_stats.get("total_amount", 0.0))
        parts.append(f"{sheet_name}: {written} ряд., {amount:,.2f} грн")

    return "RewardsAfter: колонку AO заповнено.\n" + "\n".join(parts)

def reason_from_rgb(rgb_u: str) -> str:
    if not rgb_u:
        return "інше"

    if rgb_u.endswith("FF0000"):
        return "СЗЧ"
    if rgb_u.endswith("FFA500"):
        return "зникнення"
    if rgb_u.endswith("FF9966"):
        return "арешт"
    if rgb_u.endswith("0070C0"):
        return "шпиталь"
    if rgb_u.endswith("A0A0A0") or rgb_u.endswith("606060"):
        return "не в УОС"

    return "інше"

# ===================== EXCLUDE (A contains keywords in current month) =====================

EXCLUDE_KEYWORDS = ["сзч", "смерть", "заг"]
_ISO_DATE_RE = re.compile(r"\b(\d{4})-(\d{2})-(\d{2})\b", re.IGNORECASE)

def should_exclude_from_100_70_30(a_text: str, year: int, month: int) -> bool:
    if not a_text:
        return False

    s = str(a_text).casefold()
    if not any(k in s for k in EXCLUDE_KEYWORDS):
        return False

    for m0 in _ISO_DATE_RE.finditer(s):
        yy = int(m0.group(1))
        mm = int(m0.group(2))
        if yy == year and mm == month:
            return True

    return False

# ===================== CLEAR OUTPUT SHEETS =====================

def clear_output_sheets(wb, sheets: list[str], start_row: int = 7):
    for sh in sheets:
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        max_row = ws.max_row
        for r in range(start_row, max_row + 1):
            for c in range(2, 7):  # B..F
                cell = ws.cell(row=r, column=c)
                cell.value = None
                cell.fill = NO_FILL
            if sh == "0":
                cell = ws.cell(row=r, column=8)  # H
                cell.value = None
                cell.fill = NO_FILL

def _column_width_chars(ws, col_idx: int) -> int:
    col_letter = get_column_letter(col_idx)
    width = ws.column_dimensions[col_letter].width
    try:
        width = float(width) if width is not None else OUTPUT_DEFAULT_COLUMN_WIDTH
    except (TypeError, ValueError):
        width = OUTPUT_DEFAULT_COLUMN_WIDTH
    return max(1, int(width) - 1)


def _wrapped_line_count(text: str, chars_per_line: int) -> int:
    lines = 0
    for part in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        if not part:
            lines += 1
        else:
            lines += (len(part) + chars_per_line - 1) // chars_per_line
    return max(1, lines)


def estimate_output_row_height(ws, row: int, columns: tuple[int, ...]) -> float:
    max_lines = 1
    for c in columns:
        cell = ws.cell(row=row, column=c)
        if cell.value in (None, ""):
            continue
        max_lines = max(max_lines, _wrapped_line_count(str(cell.value), _column_width_chars(ws, c)))

    height = max(OUTPUT_MIN_ROW_HEIGHT, max_lines * OUTPUT_ROW_HEIGHT_PER_LINE)
    if max_lines > 1:
        height += OUTPUT_ROW_HEIGHT_PADDING
    return min(height, OUTPUT_MAX_ROW_HEIGHT)


def apply_output_sheet_formatting(ws, start_row: int, end_row: int, columns: tuple[int, ...]) -> tuple[int, int, int]:
    if end_row < start_row:
        return 0, 0, 0

    rows_formatted = 0
    cells_formatted = 0
    row_heights_set = 0
    for r in range(start_row, end_row + 1):
        row_has_value = False
        for c in columns:
            cell = ws.cell(row=r, column=c)
            if cell.value not in (None, ""):
                row_has_value = True
            cell.alignment = OUTPUT_CENTER_WRAP_ALIGNMENT
            cells_formatted += 1

        if row_has_value:
            ws.row_dimensions[r].height = estimate_output_row_height(ws, r, columns)
            rows_formatted += 1
            row_heights_set += 1

    return rows_formatted, cells_formatted, row_heights_set

def format_output_sheets_after_write(wb, last_rows: dict[str, int], start_row: int = OUTPUT_START_ROW) -> dict[str, dict[str, int]]:
    stats: dict[str, dict[str, int]] = {}

    for sheet_name, columns in OUTPUT_SHEET_COLUMNS.items():
        if sheet_name not in wb.sheetnames:
            continue

        last_row = last_rows.get(sheet_name, start_row - 1)
        rows_formatted, cells_formatted, row_heights_set = apply_output_sheet_formatting(
            ws=wb[sheet_name],
            start_row=start_row,
            end_row=last_row,
            columns=columns,
        )
        stats[sheet_name] = {
            "last_row": last_row,
            "rows": rows_formatted,
            "cells": cells_formatted,
            "row_heights": row_heights_set,
        }

    return stats

# ===================== CORE =====================

def build_periods_for_row(
    ws100_src, ws70_src,
    row: int,
    row70: int,
    start_col_idx: int,
    last_day: int,
    year: int,
    month: int
) -> tuple[list[tuple[date,date]], list[tuple[date,date]], list[tuple[date,date]], dict[str, list[tuple[date,date]]]]:
    days100: list[int] = []
    days70: list[int] = []
    days30: list[int] = []
    days0_by_reason: dict[str, list[int]] = {}

    for day in range(1, last_day + 1):
        col = _day_to_col(start_col_idx, day)

        c100 = ws100_src.cell(row=row, column=col)
        c70  = ws70_src.cell(row=row70, column=col)

        v100 = c100.value
        v70  = c70.value

        # 100: тільки значення "1"
        if str(v100).strip() == "1":
            days100.append(day)

        # 70: тільки значення "1"
        if str(v70).strip() == "1":
            days70.append(day)

        # 30: оливковий (на будь-якому з 2 аркушів)
        rgb100 = _fill_rgb(c100)
        rgb70  = _fill_rgb(c70)
        if (rgb100.endswith(OLIVE_SUFFIX) or rgb70.endswith(OLIVE_SUFFIX)):
            days30.append(day)

        # 0: все інше (пріоритет: 100/70/30 вже забрали)
        if (day in days100) or (day in days70) or (day in days30):
            continue

        rgb = rgb100 or rgb70
        if not rgb:
            continue

        reason = reason_from_rgb(rgb)
        days0_by_reason.setdefault(reason, []).append(day)

    periods100 = _compress_days_to_periods(days100, year, month)
    periods70  = _compress_days_to_periods(days70, year, month)
    periods30  = _compress_days_to_periods(days30, year, month)

    periods0_by_reason2: dict[str, list[tuple[date,date]]] = {}
    for reason, days in days0_by_reason.items():
        periods0_by_reason2[reason] = _compress_days_to_periods(days, year, month)

    return periods100, periods70, periods30, periods0_by_reason2

def write_periods_to_sheet(
    ws_out,
    out_row_start: int,
    pos: str,
    rank: str,
    pib: str,
    periods: list[tuple[date,date]],
    reason: str | None = None
) -> int:
    r = out_row_start
    for s, e in periods:
        ws_out.cell(row=r, column=2).value = pos                 # B
        ws_out.cell(row=r, column=3).value = rank                # C
        ws_out.cell(row=r, column=4).value = pib                 # D
        ws_out.cell(row=r, column=5).value = _fmt_date_ua(s)     # E
        ws_out.cell(row=r, column=6).value = _fmt_date_ua(e)     # F
        if reason is not None:
            ws_out.cell(row=r, column=8).value = reason          # H
        r += 1
    return r

def run_rewards_after(target_xlsx: Path, start_col_idx: int, year: int, month: int, progress=None):
    last_day = calendar.monthrange(year, month)[1]

    if progress:
        progress.update(
            header="Етап 2/5: Завантаження Excel",
            detail="Відкриваю цільовий файл...",
            current=0,
            total=1,
            file_name=target_xlsx.name,
        )
    wb = load_workbook(target_xlsx)
    if progress:
        progress.update(
            header="Етап 2/5: Завантаження Excel",
            detail="Файл завантажено",
            current=1,
            total=1,
            file_name=target_xlsx.name,
        )

    # required sheets
    required = ("100 000", "70 000", "100", "70", "30", "0")
    for sh in required:
        if sh not in wb.sheetnames:
            wb.close()
            raise SystemExit(f"Нема листа '{sh}' у файлі: {target_xlsx.name}")

    ws100_src = wb["100 000"]
    ws70_src  = wb["70 000"]
    ws100_out = wb["100"]
    ws70_out  = wb["70"]
    ws30_out  = wb["30"]
    ws0_out   = wb["0"]

    # clear outputs for rerun
    if progress:
        progress.update(
            header="Етап 3/5: Очищення листів",
            detail="Очищаю попередні дані на листах 100 / 70 / 30 / 0...",
            current=0,
            total=1,
            file_name=target_xlsx.name,
        )
    clear_output_sheets(wb, sheets=["100", "70", "30", "0"], start_row=OUTPUT_START_ROW)
    if progress:
        progress.update(
            header="Етап 3/5: Очищення листів",
            detail="Листи підготовлено для нового запису",
            current=1,
            total=1,
            file_name=target_xlsx.name,
        )

    # читаємо ПІБи в порядку з "100 000" до рядка "приком"
    start_row_src = 5  # як у тебе
    prikom_cut = _find_marker_cutoff_row(ws100_src, start_row=start_row_src, marker_col_idx=4, marker="приком")
    last_row_src = ws100_src.max_row if prikom_cut is None else (prikom_cut - 1)
    prikom_cut_70 = _find_marker_cutoff_row(ws70_src, start_row=start_row_src, marker_col_idx=4, marker="приком")
    last_row_70 = ws70_src.max_row if prikom_cut_70 is None else (prikom_cut_70 - 1)

    # Align rows by PIB to avoid accidental "neighbor row" reads from 70 000.
    pib_to_rows_70 = _build_pib_row_queues(ws70_src, start_row=start_row_src, last_row=last_row_70, pib_col_idx=5)

    # output cursors
    out100_r = OUTPUT_START_ROW
    out70_r  = OUTPUT_START_ROW
    out30_r  = OUTPUT_START_ROW
    out0_r   = OUTPUT_START_ROW

    processed_people = 0
    excluded_people = 0
    wrote_100 = wrote_70 = wrote_30 = wrote_0 = 0
    used_70_matched = 0
    used_70_fallback = 0

    total_rows = max(1, last_row_src - start_row_src + 1)
    if progress:
        progress.update(
            header="Етап 4/5: Обробка рядків",
            detail="Аналізую листи 100 000 та 70 000...",
            current=0,
            total=total_rows,
            file_name="",
        )

    for row_no, r in enumerate(range(start_row_src, last_row_src + 1), start=1):
        if progress and (row_no == 1 or row_no % 10 == 0 or row_no == total_rows):
            progress.update(
                header="Етап 4/5: Обробка рядків",
                detail="Аналізую листи 100 000 та 70 000...",
                current=row_no,
                total=total_rows,
                file_name=f"Рядок {r}",
            )

        pib_raw = ws100_src.cell(row=r, column=5).value
        pib = str(pib_raw or "").strip()
        if not pib:
            continue

        processed_people += 1

        pib_key = _norm_pib(pib_raw)
        r70 = r
        q70 = pib_to_rows_70.get(pib_key)
        if q70:
            r70 = q70.popleft()
            used_70_matched += 1
        else:
            used_70_fallback += 1

        pos = str(ws100_src.cell(row=r, column=3).value or "").strip()  # C
        rank = str(ws100_src.cell(row=r, column=4).value or "").strip()  # D

        a_text = ws100_src.cell(row=r, column=1).value  # A
        is_excluded = should_exclude_from_100_70_30(a_text, year=year, month=month)
        if is_excluded:
            excluded_people += 1

        periods100, periods70, periods30, periods0_by_reason = build_periods_for_row(
            ws100_src=ws100_src,
            ws70_src=ws70_src,
            row=r,
            row70=r70,
            start_col_idx=start_col_idx,
            last_day=last_day,
            year=year,
            month=month
        )

        # 30/100/70: пропускаємо якщо "бан"
        if (not is_excluded) and periods30:
            before = out30_r
            out30_r = write_periods_to_sheet(ws30_out, out30_r, pos, rank, pib, periods30)
            wrote_30 += (out30_r - before)

        if (not is_excluded) and periods100:
            before = out100_r
            out100_r = write_periods_to_sheet(ws100_out, out100_r, pos, rank, pib, periods100)
            wrote_100 += (out100_r - before)

        if (not is_excluded) and periods70:
            before = out70_r
            out70_r = write_periods_to_sheet(ws70_out, out70_r, pos, rank, pib, periods70)
            wrote_70 += (out70_r - before)

        # 0: пишемо завжди (якщо є що)
        for reason, periods in periods0_by_reason.items():
            if not periods:
                continue
            before = out0_r
            out0_r = write_periods_to_sheet(ws0_out, out0_r, pos, rank, pib, periods, reason=reason)
            wrote_0 += (out0_r - before)

    if progress:
        progress.update(
            header="Етап 5/5: Збереження",
            detail="Форматую заповнені листи...",
            current=0,
            total=2,
            file_name=target_xlsx.name,
        )
    format_stats = format_output_sheets_after_write(
        wb,
        last_rows={
            "100": out100_r - 1,
            "70": out70_r - 1,
            "30": out30_r - 1,
            "0": out0_r - 1,
        },
        start_row=OUTPUT_START_ROW,
    )

    if progress:
        progress.update(
            header="Етап 5/5: Збереження",
            detail="Зберігаю Excel-файл...",
            current=1,
            total=2,
            file_name=target_xlsx.name,
        )
    save_target_workbook(wb, target_xlsx)
    wb.close()

    if progress:
        progress.update(
            header="Етап 5/5: Збереження",
            detail="Завершено!",
            current=2,
            total=2,
            file_name=target_xlsx.name,
        )

    log(f"[AFTER] processed_people={processed_people}")
    log(f"[AFTER] row align 70k: matched_by_pib={used_70_matched}, fallback_same_row={used_70_fallback}")
    log(f"[AFTER] excluded_from_100_70_30={excluded_people}")
    log(f"[AFTER] rows_written: 100={wrote_100}, 70={wrote_70}, 30={wrote_30}, 0={wrote_0}")
    log(f"[AFTER] formatting_stats={format_stats}")
    log("[AFTER] row_autofit_via_openpyxl=True")

# ===================== MAIN =====================

def main():
    ui_root = None
    progress = None
    try:
        run_start()

        target_xlsx, start_col_idx, year, month, money_mode = ask_all_inputs_after()

        log(f"[INPUT] TARGET file : {target_xlsx}")
        log(f"[PARAMS] year={year}, month={month}, start_col_idx={start_col_idx}, money_mode={money_mode}")

        ui_root = Tk()
        ui_root.withdraw()
        install_frozen_executable_icon(ui_root)
        install_dark_title_bar(ui_root)
        ui_root.attributes("-topmost", True)

        progress = ProgressWindow(
            ui_root,
            subtitle=(
                "Перерахунок зарплат у колонці AO виконується..."
                if money_mode
                else "Заповнення листів 100 / 70 / 30 / 0 виконується..."
            ),
        )
        if money_mode:
            stats = run_rewards_money_with_progress(
                target_xlsx=target_xlsx,
                start_col_idx=start_col_idx,
                year=year,
                month=month,
                progress=progress,
            )
            for sheet_name, sheet_stats in stats.items():
                log(f"[MONEY] {sheet_name}: {sheet_stats}")
            success_detail = "RewardsAfter: колонку AO заповнено."
            success_message = format_money_success_message(stats)
        else:
            run_rewards_after_with_progress(
                target_xlsx=target_xlsx,
                start_col_idx=start_col_idx,
                year=year,
                month=month,
                progress=progress,
            )
            success_detail = "RewardsAfter: листи 100 / 70 / 30 / 0 заповнені."
            success_message = "RewardsAfter: листи 100/70/30/0 заповнені."

        run_end()
        progress.show_success_then_close(
            success_detail,
            file_name=target_xlsx.name,
        )
        progress = None

        if ui_root is not None and _widget_exists(ui_root):
            ui_root.attributes("-topmost", True)
            messagebox.showinfo("Готово!", success_message, parent=ui_root)
    except SystemExit as exc:
        if str(exc):
            log(str(exc))
    except Exception as exc:
        error_parent = None
        if progress is not None and _widget_exists(progress.dialog):
            error_parent = progress.dialog
        elif ui_root is not None and _widget_exists(ui_root):
            error_parent = ui_root

        try:
            if error_parent is not None:
                error_parent.lift()
                error_parent.attributes("-topmost", True)
                error_parent.update()
                messagebox.showerror(getattr(exc, "title", "Помилка"), str(exc), parent=error_parent)
            else:
                root = Tk()
                root.withdraw()
                install_frozen_executable_icon(root)
                install_dark_title_bar(root)
                root.attributes("-topmost", True)
                messagebox.showerror(getattr(exc, "title", "Помилка"), str(exc), parent=root)
                root.destroy()
        except Exception:
            print(f"Помилка: {exc}")
    finally:
        if progress is not None:
            progress.close()
        if ui_root is not None and _widget_exists(ui_root):
            try:
                ui_root.destroy()
            except Exception:
                pass

if __name__ == "__main__":
    main()
