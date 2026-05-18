#!python3.11
from __future__ import annotations

from collections import deque
from datetime import datetime
from pathlib import Path
from tkinter import Tk, filedialog, messagebox, StringVar, ttk
from zipfile import BadZipFile, ZipFile
import calendar
import os
import queue
import re
import shutil
import sys
import threading
import time
import tkinter as tk
import tkinter.font as tkfont

from openpyxl import load_workbook as _openpyxl_load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string, get_column_letter
from send2trash import send2trash


APP_NAME = "RewardsGroshi"

REQUIRED_SHEET = "100 000"
OPTIONAL_SHEET = "упр"
TARGET_SHEETS = (REQUIRED_SHEET, OPTIONAL_SHEET)

START_COL_LETTER = "F"
START_COL_IDX = column_index_from_string(START_COL_LETTER)
HEADER_ROWS_TO_SKIP = 5
DEFAULT_DATA_START_ROW = HEADER_ROWS_TO_SKIP + 1
PIB_COL_IDX = column_index_from_string("E")
MONEY_COL_LETTER = "AN"
MONEY_COL_IDX = column_index_from_string(MONEY_COL_LETTER)

OLIVE_MONTHLY_RATE = 30000.0
GREEN_MONTHLY_RATE = 100000.0
OLIVE_SUFFIX = "C4D79B"
GREEN_SUFFIXES = ("00B050", "006100")

HEADER_PERIOD_RE = re.compile(r"\b(20\d{2})-(0[1-9]|1[0-2])\b")
HEADER_PERIOD_SCAN_SHEETS = (REQUIRED_SHEET, OPTIONAL_SHEET, "70 000", "100", "70", "30", "0")

# Excel accounting-like hryvnia format. Values stay numeric; Excel displays two decimals.
ACCOUNTING_UAH_FORMAT = '_-* #,##0.00\\ "грн"_-;-* #,##0.00\\ "грн"_-;_-* "-"??\\ "грн"_-;_-@_-'
MONEY_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=False)


def install_frozen_executable_icon(root, retry_ms: int = 250) -> None:
    if sys.platform != "win32" or not getattr(sys, "frozen", False):
        return

    _set_frozen_executable_icon(root)
    try:
        root.after(retry_ms, lambda: _set_frozen_executable_icon(root))
    except Exception:
        pass


def _set_frozen_executable_icon(root) -> None:
    if sys.platform != "win32" or not getattr(sys, "frozen", False):
        return

    try:
        import ctypes
        from ctypes import wintypes

        wm_seticon = 0x0080
        icon_small = 0
        icon_big = 1
        gclp_hicon = -14
        gclp_hiconsm = -34

        shell32 = ctypes.WinDLL("shell32", use_last_error=True)
        user32 = ctypes.WinDLL("user32", use_last_error=True)
        long_ptr = ctypes.c_longlong if ctypes.sizeof(ctypes.c_void_p) == 8 else ctypes.c_long

        shell32.ExtractIconExW.argtypes = (
            wintypes.LPCWSTR,
            ctypes.c_int,
            ctypes.POINTER(wintypes.HICON),
            ctypes.POINTER(wintypes.HICON),
            wintypes.UINT,
        )
        shell32.ExtractIconExW.restype = wintypes.UINT
        user32.SendMessageW.argtypes = (
            wintypes.HWND,
            wintypes.UINT,
            wintypes.WPARAM,
            wintypes.LPARAM,
        )
        user32.SendMessageW.restype = long_ptr
        user32.GetParent.argtypes = (wintypes.HWND,)
        user32.GetParent.restype = wintypes.HWND

        try:
            set_class_icon = user32.SetClassLongPtrW
        except AttributeError:
            set_class_icon = user32.SetClassLongW
        set_class_icon.argtypes = (wintypes.HWND, ctypes.c_int, long_ptr)
        set_class_icon.restype = long_ptr

        icon_handles = getattr(root, "_frozen_executable_icon_handles", None)
        if icon_handles is None:
            large_icons = (wintypes.HICON * 1)()
            small_icons = (wintypes.HICON * 1)()
            icon_count = shell32.ExtractIconExW(sys.executable, 0, large_icons, small_icons, 1)
            if icon_count <= 0:
                return
            icon_handles = (small_icons[0], large_icons[0])
            root._frozen_executable_icon_handles = icon_handles
        small_icon, large_icon = icon_handles

        root.update_idletasks()
        hwnds = []
        for raw_hwnd in (root.winfo_id(), root.wm_frame()):
            try:
                hwnd = int(str(raw_hwnd), 0)
            except (TypeError, ValueError):
                continue
            while hwnd and hwnd not in hwnds:
                hwnds.append(hwnd)
                hwnd = user32.GetParent(hwnd)

        for hwnd in hwnds:
            if small_icon:
                user32.SendMessageW(hwnd, wm_seticon, icon_small, small_icon)
                set_class_icon(hwnd, gclp_hiconsm, small_icon)
            if large_icon:
                user32.SendMessageW(hwnd, wm_seticon, icon_big, large_icon)
                set_class_icon(hwnd, gclp_hicon, large_icon)
    except Exception:
        pass


def install_dark_title_bar(window, retry_ms: int = 80) -> None:
    if sys.platform != "win32":
        return

    def _apply() -> None:
        try:
            import ctypes

            hwnd = ctypes.windll.user32.GetParent(window.winfo_id())
            value = ctypes.c_int(1)
            for attr in (20, 19):
                ctypes.windll.dwmapi.DwmSetWindowAttribute(
                    hwnd,
                    attr,
                    ctypes.byref(value),
                    ctypes.sizeof(value),
                )
        except Exception:
            pass

    try:
        window.update_idletasks()
        _apply()
        window.after(retry_ms, _apply)
    except Exception:
        pass


def _widget_exists(widget) -> bool:
    try:
        return bool(widget.winfo_exists())
    except Exception:
        return False


def _ts() -> str:
    return datetime.now().strftime("%H:%M:%S")


def log(message: str) -> None:
    print(f"[{_ts()}] {message}")


_RUN_T0 = None


def run_start() -> None:
    global _RUN_T0
    _RUN_T0 = time.perf_counter()
    log(f"=== START {APP_NAME} ===")


def run_end() -> None:
    if _RUN_T0 is None:
        log(f"=== END {APP_NAME} (no timer) ===")
        return
    dt = time.perf_counter() - _RUN_T0
    log(f"=== END {APP_NAME} === total={dt:.2f}s ({dt / 60:.2f} min)")


def load_workbook(filename, *args, **kwargs):
    try:
        suffix = Path(filename).suffix.lower()
    except TypeError:
        suffix = ""

    if suffix == ".xlsm":
        kwargs.setdefault("keep_vba", True)
    kwargs.setdefault("keep_links", True)
    return _openpyxl_load_workbook(filename, *args, **kwargs)


def _assert_valid_xlsx_package(path: Path) -> None:
    try:
        with ZipFile(path) as zf:
            bad_member = zf.testzip()
    except BadZipFile as exc:
        raise ValueError(f"Збережений Excel-файл має пошкоджений ZIP-контейнер: {path}") from exc

    if bad_member:
        raise ValueError(f"Збережений Excel-файл має пошкоджену частину ZIP: {bad_member}")


def save_workbook_safely(wb, target_xlsx: Path) -> None:
    target_xlsx = Path(target_xlsx)
    tmp_path = target_xlsx.with_name(f".{target_xlsx.stem}.{os.getpid()}.tmp{target_xlsx.suffix}")

    try:
        if tmp_path.exists():
            tmp_path.unlink()
        wb.save(tmp_path)
        _assert_valid_xlsx_package(tmp_path)
        tmp_path.replace(target_xlsx)
    finally:
        try:
            if tmp_path.exists():
                tmp_path.unlink()
        except Exception:
            pass


def format_path_for_entry_display(path: str) -> str:
    path = path.strip()
    if not path:
        return ""

    normalized = Path(path)
    drive_prefix = ""
    if normalized.drive and len(normalized.drive) == 2 and normalized.drive[1] == ":":
        drive_prefix = f"{normalized.drive}/"

    name = normalized.name
    parent_name = normalized.parent.name
    if parent_name and name:
        return f"{drive_prefix}.../{parent_name}/{name}"
    if name:
        return f"{drive_prefix}{name}" if drive_prefix else f".../{name}"
    if drive_prefix:
        return drive_prefix
    return path.replace("\\", "/")


def fit_path_for_entry_display(path: str, entry) -> str:
    display_path = format_path_for_entry_display(path)
    if not display_path:
        return ""

    entry.update_idletasks()
    font_name = entry.cget("font") or "TkDefaultFont"
    try:
        font = tkfont.nametofont(font_name)
    except Exception:
        return display_path

    available_width = max(120, entry.winfo_width() - 14)
    if font.measure(display_path) <= available_width:
        return display_path

    parts = display_path.split("/")
    if len(parts) >= 3:
        prefix = "/".join(parts[:-1]) + "/..."
        shortened_name = parts[-1]
        while len(shortened_name) > 1 and font.measure(prefix + shortened_name) > available_width:
            shortened_name = shortened_name[1:]
        return prefix + shortened_name

    ellipsis = "..."
    shortened_path = display_path
    while len(shortened_path) > 1 and font.measure(ellipsis + shortened_path) > available_width:
        shortened_path = shortened_path[1:]
    return ellipsis + shortened_path


def refresh_path_entry_display(source_var: StringVar, display_var: StringVar, entry) -> None:
    if not _widget_exists(entry):
        return
    display_var.set(fit_path_for_entry_display(source_var.get(), entry))


def _parse_header_year_month(value) -> tuple[int, int] | None:
    if value is None:
        return None
    match = HEADER_PERIOD_RE.search(str(value))
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


def detect_template_period(
    target_xlsx: Path,
    preferred_sheets: tuple[str, ...] = HEADER_PERIOD_SCAN_SHEETS,
) -> tuple[int, int]:
    wb = load_workbook(target_xlsx, read_only=True, data_only=True)
    try:
        ordered_sheet_names: list[str] = []
        seen: set[str] = set()
        for sheet_name in (*preferred_sheets, *wb.sheetnames):
            if sheet_name in wb.sheetnames and sheet_name not in seen:
                ordered_sheet_names.append(sheet_name)
                seen.add(sheet_name)

        for sheet_name in ordered_sheet_names:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=5, values_only=True):
                for value in row:
                    parsed = _parse_header_year_month(value)
                    if parsed is not None:
                        return parsed
    finally:
        wb.close()

    raise ValueError("Не знайшов формат YYYY-MM у заголовку шаблону (A1:E3).")


def detect_template_context(target_xlsx: Path) -> tuple[int, int, int]:
    year, month = detect_template_period(target_xlsx)
    return START_COL_IDX, year, month


def configure_launch_styles(root):
    colors = {
        "window": "#F4F7FB",
        "panel": "#FFFFFF",
        "border": "#D6DEE8",
        "text": "#15202B",
        "muted": "#617084",
        "accent": "#0F766E",
        "accent_dark": "#0B5D56",
        "header": "#103C3A",
        "header_badge": "#0F766E",
    }

    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except Exception:
        pass

    try:
        style.configure("LaunchRoot.TFrame", background=colors["window"])
        style.configure("LaunchPanel.TFrame", background=colors["panel"])
        style.configure(
            "LaunchSection.TLabel",
            background=colors["panel"],
            foreground=colors["text"],
            font=("Segoe UI", 12, "bold"),
        )
        style.configure(
            "LaunchField.TLabel",
            background=colors["panel"],
            foreground=colors["text"],
            font=("Segoe UI", 10, "bold"),
        )
        style.configure(
            "LaunchMuted.TLabel",
            background=colors["panel"],
            foreground=colors["muted"],
            font=("Segoe UI", 9),
        )
        style.configure(
            "ProgressHeader.TLabel",
            background=colors["panel"],
            foreground=colors["text"],
            font=("Segoe UI", 12, "bold"),
        )
        style.configure(
            "ProgressBody.TLabel",
            background=colors["panel"],
            foreground=colors["text"],
            font=("Segoe UI", 10),
        )
        style.configure(
            "ProgressFile.TLabel",
            background=colors["panel"],
            foreground=colors["muted"],
            font=("Segoe UI", 9),
        )
        style.configure(
            "ProgressCount.TLabel",
            background=colors["panel"],
            foreground=colors["accent_dark"],
            font=("Segoe UI", 10, "bold"),
        )
        style.configure(
            "Launch.Horizontal.TProgressbar",
            background=colors["header_badge"],
            troughcolor="#E8EEF5",
            bordercolor=colors["border"],
            lightcolor=colors["header_badge"],
            darkcolor=colors["header_badge"],
            thickness=12,
        )
        style.configure(
            "Launch.TEntry",
            fieldbackground="#FFFFFF",
            foreground=colors["text"],
            bordercolor=colors["border"],
            lightcolor=colors["border"],
            darkcolor=colors["border"],
            padding=(8, 6),
        )
        style.map(
            "Launch.TEntry",
            fieldbackground=[("readonly", "#FFFFFF")],
            foreground=[("readonly", colors["text"])],
        )
        for style_name in ("LaunchBrowse.TButton", "LaunchPrimary.TButton", "LaunchHero.TButton"):
            style.configure(
                style_name,
                font=("Segoe UI", 10, "bold"),
                padding=(18, 10),
                foreground="#FFFFFF",
                background="#053D39",
                bordercolor=colors["accent_dark"],
                lightcolor=colors["accent_dark"],
                darkcolor=colors["accent_dark"],
                focuscolor=colors["accent_dark"],
            )
            style.map(
                style_name,
                background=[("pressed", colors["accent_dark"]), ("active", colors["accent_dark"])],
                foreground=[("pressed", "#FFFFFF"), ("active", "#FFFFFF")],
            )
    except Exception:
        pass

    return colors


class ProgressWindow:
    def __init__(self, owner: Tk):
        self.close_requested = False
        self.dialog = tk.Toplevel(owner)
        self.dialog.withdraw()
        install_frozen_executable_icon(self.dialog)
        self.dialog.title(f"{APP_NAME} - Обробка")
        self.dialog.resizable(False, False)
        self.dialog.attributes("-topmost", True)
        self.dialog.protocol("WM_DELETE_WINDOW", self.request_close)
        colors = configure_launch_styles(self.dialog)
        self.dialog.configure(bg=colors["window"])
        install_dark_title_bar(self.dialog)

        container = ttk.Frame(self.dialog, style="LaunchRoot.TFrame")
        container.pack(fill="both", expand=True)

        header = tk.Frame(container, bg=colors["header"], padx=22, pady=10)
        header.pack(fill="x")
        header.grid_columnconfigure(0, weight=1)
        tk.Label(
            header,
            text=APP_NAME,
            bg=colors["header"],
            fg="#FFFFFF",
            font=("Segoe UI Semibold", 15),
            anchor="w",
        ).grid(row=0, column=0, sticky="w")
        tk.Label(
            header,
            text="Розрахунок зарплат у колонці AN...",
            bg=colors["header"],
            fg="#D7FBF5",
            font=("Segoe UI", 9),
            anchor="w",
        ).grid(row=1, column=0, sticky="w", pady=(3, 4))
        badge_shell = tk.Frame(header, bg=colors["accent_dark"], padx=1, pady=1)
        badge = tk.Label(
            badge_shell,
            text="ОБРОБКА",
            bg="#053D39",
            fg="#FFFFFF",
            font=("Segoe UI", 12, "bold"),
            padx=10,
            pady=5,
        )
        badge.pack()
        badge_shell.place(relx=1.0, x=5, y=5, anchor="ne")

        body = ttk.Frame(container, style="LaunchRoot.TFrame", padding=(18, 16, 18, 18))
        body.pack(fill="both", expand=True)
        panel_shell = tk.Frame(body, bg=colors["border"], padx=1, pady=1)
        panel_shell.pack(fill="both", expand=True)
        frame = ttk.Frame(panel_shell, style="LaunchPanel.TFrame", padding=(16, 14, 16, 12))
        frame.pack(fill="both", expand=True)
        frame.grid_columnconfigure(0, weight=1)

        self.header_var = tk.StringVar(value="Підготовка...")
        self.detail_var = tk.StringVar(value="Будь ласка, зачекайте")
        self.file_var = tk.StringVar(value="")
        self.count_var = tk.StringVar(value="0 / 0")

        ttk.Label(frame, textvariable=self.header_var, style="ProgressHeader.TLabel").grid(
            row=0, column=0, padx=(0, 12), sticky="w"
        )
        ttk.Label(frame, textvariable=self.count_var, style="ProgressCount.TLabel").grid(
            row=0, column=1, sticky="e"
        )
        ttk.Label(frame, textvariable=self.detail_var, style="ProgressBody.TLabel", wraplength=500).grid(
            row=1, column=0, columnspan=2, pady=(8, 6), sticky="w"
        )
        self.progress = ttk.Progressbar(
            frame,
            orient="horizontal",
            mode="determinate",
            maximum=100,
            style="Launch.Horizontal.TProgressbar",
        )
        self.progress.grid(row=2, column=0, columnspan=2, pady=(6, 0), sticky="we")
        ttk.Label(frame, textvariable=self.file_var, style="ProgressFile.TLabel", wraplength=500).grid(
            row=3, column=0, columnspan=2, pady=(8, 0), sticky="w"
        )

        self.dialog.update_idletasks()
        width = max(560, self.dialog.winfo_reqwidth())
        height = max(238, self.dialog.winfo_reqheight())
        self.dialog.geometry(f"{width}x{height}")
        self.dialog.deiconify()
        self.dialog.lift()
        self.refresh()

    def update(self, header=None, detail=None, current=None, total=None, file_name=None):
        try:
            if self.close_requested or not _widget_exists(self.dialog):
                return
            if header is not None:
                self.header_var.set(header)
            if detail is not None:
                self.detail_var.set(detail)
            if file_name is not None:
                self.file_var.set(file_name)
            if current is not None and total is not None:
                safe_total = total if total > 0 else 1
                pct = max(0.0, min(100.0, (current / safe_total) * 100.0))
                self.progress["value"] = pct
                self.count_var.set(f"{current} / {total}")
            self.refresh()
        except tk.TclError:
            return

    def refresh(self):
        try:
            if self.close_requested or not _widget_exists(self.dialog):
                return
            self.dialog.update_idletasks()
            self.dialog.update()
        except Exception:
            pass

    def close(self):
        self.close_requested = True
        try:
            if _widget_exists(self.dialog):
                self.dialog.destroy()
        except Exception:
            pass

    def request_close(self):
        self.close()

    def show_success_then_close(self, detail: str, file_name: str = "", delay_ms: int = 1300):
        if self.close_requested or not _widget_exists(self.dialog):
            return
        self.update(
            header="Готово",
            detail=detail,
            current=1,
            total=1,
            file_name=file_name,
        )
        deadline = time.monotonic() + delay_ms / 1000.0
        while time.monotonic() < deadline:
            if not _widget_exists(self.dialog):
                return
            self.refresh()
            time.sleep(0.05)
        self.close()


class ProgressUpdateProxy:
    def __init__(self, updates: "queue.Queue[dict[str, object]]"):
        self.updates = updates

    def update(self, header=None, detail=None, current=None, total=None, file_name=None):
        self.updates.put(
            {
                "header": header,
                "detail": detail,
                "current": current,
                "total": total,
                "file_name": file_name,
            }
        )


MIN_PROGRESS_STAGE_SECONDS = 1.1
PROGRESS_POLL_SECONDS = 0.03


class ProgressDisplayGate:
    def __init__(self, progress: ProgressWindow, min_stage_seconds: float = MIN_PROGRESS_STAGE_SECONDS):
        self.progress = progress
        self.min_stage_seconds = min_stage_seconds
        self.pending: deque[dict[str, object]] = deque()
        self.active_header = None
        self.active_since = 0.0

    def drain_updates(self, updates: "queue.Queue[dict[str, object]]") -> None:
        while True:
            try:
                self.pending.append(updates.get_nowait())
            except queue.Empty:
                return

    def has_pending(self) -> bool:
        return bool(self.pending)

    def apply_ready(self) -> bool:
        if not self.pending:
            return False

        payload = self.pending[0]
        next_header = payload.get("header") or self.active_header
        now = time.monotonic()

        if (
            self.active_header is not None
            and next_header != self.active_header
            and now - self.active_since < self.min_stage_seconds
        ):
            return False

        payload = self.pending.popleft()
        while self.pending:
            following_payload = self.pending[0]
            following_header = following_payload.get("header") or next_header
            if following_header != next_header:
                break
            payload = self.pending.popleft()

        if next_header != self.active_header:
            self.active_header = next_header
            self.active_since = time.monotonic()
        self.progress.update(**payload)
        return True


def ask_all_inputs_groshi() -> tuple[Path, int, int, int]:
    root = Tk()
    root.withdraw()
    install_frozen_executable_icon(root)
    root.title(f"НАЛАШТУВАННЯ: {APP_NAME}")
    root.attributes("-topmost", True)
    root.resizable(False, False)
    root.protocol("WM_DELETE_WINDOW", root.destroy)
    colors = configure_launch_styles(root)
    root.configure(bg=colors["window"])
    install_dark_title_bar(root)

    file_var = StringVar(value="")
    file_display_var = StringVar(value="")
    entry_width = 42

    def browse_file():
        root.lift()
        root.update()
        path = filedialog.askopenfilename(
            title="Виберіть ЦІЛЬ",
            parent=root,
            filetypes=[("Excel files", "*.xlsx;*.xlsm"), ("All files", "*.*")],
        )
        if path:
            file_var.set(path)
            refresh_path_entry_display(file_var, file_display_var, file_entry)
        root.lift()
        root.update()

    surface = ttk.Frame(root, style="LaunchRoot.TFrame")
    surface.grid(row=0, column=0, sticky="nsew")
    surface.grid_columnconfigure(0, weight=1)

    header = tk.Frame(surface, bg=colors["header"], padx=22, pady=12)
    header.grid(row=0, column=0, sticky="we")
    header.grid_columnconfigure(0, weight=1)
    tk.Label(
        header,
        text=APP_NAME,
        bg=colors["header"],
        fg="#FFFFFF",
        font=("Segoe UI Semibold", 16),
        anchor="w",
    ).grid(row=0, column=0, sticky="w")
    tk.Label(
        header,
        text="Розрахунок зарплат за кольорами у колонку AN.",
        bg=colors["header"],
        fg="#D7FBF5",
        font=("Segoe UI", 9),
        anchor="w",
    ).grid(row=1, column=0, sticky="w", pady=(4, 6))
    badge_shell = tk.Frame(header, bg=colors["accent_dark"], padx=1, pady=1)
    badge = tk.Label(
        badge_shell,
        text="НАЛАШТУВАННЯ",
        bg="#053D39",
        fg="#FFFFFF",
        font=("Segoe UI", 12, "bold"),
        padx=10,
        pady=5,
    )
    badge.pack()
    badge_shell.place(relx=1.0, x=5, y=5, anchor="ne")

    body = ttk.Frame(surface, style="LaunchRoot.TFrame", padding=(18, 16, 18, 18))
    body.grid(row=1, column=0, sticky="nsew")
    panel_shell = tk.Frame(body, bg=colors["border"], padx=1, pady=1)
    panel_shell.grid(row=0, column=0, sticky="we")
    container = ttk.Frame(panel_shell, style="LaunchPanel.TFrame", padding=(16, 14, 16, 12))
    container.pack(fill="both", expand=True)
    container.grid_columnconfigure(1, weight=1)

    ttk.Label(container, text="Ціль", style="LaunchSection.TLabel").grid(
        row=0, column=0, columnspan=3, sticky="w"
    )
    ttk.Label(container, text="ЦІЛЬ (.xlsx):", style="LaunchField.TLabel").grid(
        row=1, column=0, sticky="e", padx=8, pady=6
    )
    file_entry = ttk.Entry(
        container,
        textvariable=file_display_var,
        width=entry_width,
        state="readonly",
        style="Launch.TEntry",
    )
    file_entry.grid(row=1, column=1, sticky="we", padx=8, pady=6)
    ttk.Button(container, text="ОБРАТИ", width=12, command=browse_file, style="LaunchBrowse.TButton").grid(
        row=1, column=2, sticky="w", padx=8, pady=6
    )
    file_var.trace_add("write", lambda *_args: refresh_path_entry_display(file_var, file_display_var, file_entry))
    file_entry.bind("<Configure>", lambda _event: refresh_path_entry_display(file_var, file_display_var, file_entry))

    footer_line = tk.Frame(container, bg=colors["border"], height=1)
    footer_line.grid(row=2, column=0, columnspan=3, sticky="we", padx=8, pady=(12, 14))

    result = {"ok": False, "data": None}

    def validate_and_close():
        try:
            path_text = file_var.get().strip()
            if not path_text:
                raise ValueError("Не вибрано файл-ціль.")

            target = Path(path_text)
            if not target.exists():
                raise ValueError("Файл не існує.")

            start_col_idx, year, month = detect_template_context(target)
            result["data"] = (target, start_col_idx, year, month)
            result["ok"] = True
            root.destroy()
        except Exception as exc:
            messagebox.showerror("Помилка", str(exc), parent=root)

    ttk.Button(
        container,
        text="ЗАПУСТИТИ ОБРОБКУ",
        command=validate_and_close,
        style="LaunchHero.TButton",
    ).grid(row=3, column=0, columnspan=3, sticky="we", padx=8, pady=(0, 4))

    root.update_idletasks()
    root.geometry(f"{root.winfo_reqwidth()}x{root.winfo_reqheight()}")
    root.deiconify()
    root.lift()
    root.grab_set()
    root.focus_force()
    root.mainloop()

    if not result["ok"]:
        raise SystemExit("Скасовано користувачем.")

    return result["data"]


def backup_excel_to_trash(xlsx_path: Path, progress=None) -> None:
    timestamp = datetime.now().strftime("%Y.%m.%d %H.%M")
    backup_name = f"{xlsx_path.stem} - {timestamp}{xlsx_path.suffix}"
    backup_path = xlsx_path.with_name(backup_name)

    if progress:
        progress.update(
            header="Етап 1/4: Резервна копія",
            detail="Створюю backup цільового Excel...",
            current=0,
            total=1,
            file_name=xlsx_path.name,
        )

    shutil.copy2(xlsx_path, backup_path)
    send2trash(str(backup_path))

    if progress:
        progress.update(
            header="Етап 1/4: Резервна копія",
            detail="Backup створено та переміщено в кошик",
            current=1,
            total=1,
            file_name=backup_path.name,
        )


def _fill_rgb(cell) -> str:
    try:
        fill = cell.fill
        if not fill or fill.patternType is None:
            return ""
        fg = getattr(fill, "fgColor", None)
        rgb = getattr(fg, "rgb", None) if fg else None
        return str(rgb).upper() if rgb else ""
    except Exception:
        return ""


def _day_to_col(start_col_idx: int, day: int) -> int:
    return start_col_idx + (day - 1)


def _is_olive_fill(cell) -> bool:
    return _fill_rgb(cell).endswith(OLIVE_SUFFIX)


def _is_green_fill(cell) -> bool:
    rgb = _fill_rgb(cell)
    return any(rgb.endswith(suffix) for suffix in GREEN_SUFFIXES)


def _is_person_pib(value) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    return text.casefold() != "піб"


def find_last_text_row(ws, col_idx: int, start_row: int) -> int:
    last = start_row - 1
    for row in range(start_row, ws.max_row + 1):
        value = ws.cell(row=row, column=col_idx).value
        if value is not None and str(value).strip() != "":
            last = row
    return last


def find_pib_header_row(ws) -> int | None:
    for row in range(1, min(ws.max_row, 12) + 1):
        value = ws.cell(row=row, column=PIB_COL_IDX).value
        if str(value or "").strip().casefold() == "піб":
            return row
    return None


def detect_data_start_row(ws) -> int:
    header_row = find_pib_header_row(ws)
    if header_row is None:
        return DEFAULT_DATA_START_ROW

    search_to = max(DEFAULT_DATA_START_ROW, header_row + 1)
    for row in range(header_row + 1, min(ws.max_row, search_to) + 1):
        if _is_person_pib(ws.cell(row=row, column=PIB_COL_IDX).value):
            return row

    return DEFAULT_DATA_START_ROW


def clear_money_column_and_set_format(ws) -> int:
    cleared = 0
    col_letter = get_column_letter(MONEY_COL_IDX)

    try:
        ws.column_dimensions[col_letter].number_format = ACCOUNTING_UAH_FORMAT
    except Exception:
        pass

    for row in range(1, max(ws.max_row, 1) + 1):
        cell = ws.cell(row=row, column=MONEY_COL_IDX)
        if isinstance(cell, MergedCell):
            continue
        if cell.value not in (None, ""):
            cleared += 1
        cell.value = None
        cell.number_format = ACCOUNTING_UAH_FORMAT

    return cleared


def calculate_row_salary(ws, row: int, start_col_idx: int, last_day: int) -> tuple[float, int, int]:
    olive_days = 0
    green_days = 0

    for day in range(1, last_day + 1):
        cell = ws.cell(row=row, column=_day_to_col(start_col_idx, day))
        if _is_olive_fill(cell):
            olive_days += 1
        elif _is_green_fill(cell):
            green_days += 1

    olive_daily = OLIVE_MONTHLY_RATE / last_day
    green_daily = GREEN_MONTHLY_RATE / last_day
    amount = round((olive_days * olive_daily) + (green_days * green_daily), 2)
    return amount, olive_days, green_days


def process_salary_sheet(ws, start_col_idx: int, last_day: int, progress=None, sheet_no: int = 1, total_sheets: int = 1) -> dict[str, object]:
    cleared_values = clear_money_column_and_set_format(ws)
    data_start_row = detect_data_start_row(ws)
    last_row = find_last_text_row(ws, PIB_COL_IDX, data_start_row)

    stats: dict[str, object] = {
        "sheet": ws.title,
        "data_start_row": data_start_row,
        "last_row": last_row,
        "cleared_values": cleared_values,
        "processed_rows": 0,
        "written_rows": 0,
        "olive_days": 0,
        "green_days": 0,
        "total_amount": 0.0,
    }

    if last_row < data_start_row:
        return stats

    total_rows = last_row - data_start_row + 1
    for idx, row in enumerate(range(data_start_row, last_row + 1), start=1):
        if progress and (idx == 1 or idx % 25 == 0 or idx == total_rows):
            progress.update(
                header=f"Етап 3/4: Аркуш {sheet_no}/{total_sheets}",
                detail=f"Рахую зарплати на аркуші '{ws.title}'...",
                current=idx,
                total=total_rows,
                file_name=f"Рядок {row}",
            )

        pib_value = ws.cell(row=row, column=PIB_COL_IDX).value
        if not _is_person_pib(pib_value):
            continue

        amount, olive_days, green_days = calculate_row_salary(ws, row, start_col_idx, last_day)
        target_cell = ws.cell(row=row, column=MONEY_COL_IDX)
        if isinstance(target_cell, MergedCell):
            raise ValueError(f"Клітинка {ws.title}!{target_cell.coordinate} об'єднана, не можу записати суму.")
        target_cell.value = amount
        target_cell.number_format = ACCOUNTING_UAH_FORMAT
        target_cell.alignment = MONEY_ALIGNMENT

        stats["processed_rows"] = int(stats["processed_rows"]) + 1
        stats["written_rows"] = int(stats["written_rows"]) + 1
        stats["olive_days"] = int(stats["olive_days"]) + olive_days
        stats["green_days"] = int(stats["green_days"]) + green_days
        stats["total_amount"] = float(stats["total_amount"]) + amount

    stats["total_amount"] = round(float(stats["total_amount"]), 2)
    return stats


def run_rewards_groshi(target_xlsx: Path, start_col_idx: int, year: int, month: int, progress=None) -> dict[str, dict[str, object]]:
    last_day = calendar.monthrange(year, month)[1]

    if progress:
        progress.update(
            header="Етап 2/4: Завантаження Excel",
            detail="Відкриваю цільовий файл...",
            current=0,
            total=1,
            file_name=target_xlsx.name,
        )

    wb = load_workbook(target_xlsx)
    try:
        if REQUIRED_SHEET not in wb.sheetnames:
            raise ValueError(f"Нема аркуша '{REQUIRED_SHEET}' у файлі: {target_xlsx.name}")

        sheet_names = [sheet_name for sheet_name in TARGET_SHEETS if sheet_name in wb.sheetnames]
        if progress:
            progress.update(
                header="Етап 2/4: Завантаження Excel",
                detail=f"Період: {year}-{month:02d}, днів: {last_day}",
                current=1,
                total=1,
                file_name=", ".join(sheet_names),
            )

        stats: dict[str, dict[str, object]] = {}
        for sheet_no, sheet_name in enumerate(sheet_names, start=1):
            stats[sheet_name] = process_salary_sheet(
                wb[sheet_name],
                start_col_idx=start_col_idx,
                last_day=last_day,
                progress=progress,
                sheet_no=sheet_no,
                total_sheets=len(sheet_names),
            )

        if progress:
            progress.update(
                header="Етап 4/4: Збереження",
                detail="Зберігаю Excel-файл...",
                current=0,
                total=1,
                file_name=target_xlsx.name,
            )

        save_workbook_safely(wb, target_xlsx)

        if progress:
            progress.update(
                header="Етап 4/4: Збереження",
                detail="Файл збережено",
                current=1,
                total=1,
                file_name=target_xlsx.name,
            )

        return stats
    finally:
        wb.close()


def run_rewards_groshi_with_progress(
    target_xlsx: Path,
    start_col_idx: int,
    year: int,
    month: int,
    progress: ProgressWindow,
) -> dict[str, dict[str, object]]:
    updates: "queue.Queue[dict[str, object]]" = queue.Queue()
    result = {"error": None, "stats": None}
    progress_proxy = ProgressUpdateProxy(updates)
    display_gate = ProgressDisplayGate(progress)

    def worker() -> None:
        try:
            backup_excel_to_trash(target_xlsx, progress=progress_proxy)
            result["stats"] = run_rewards_groshi(
                target_xlsx=target_xlsx,
                start_col_idx=start_col_idx,
                year=year,
                month=month,
                progress=progress_proxy,
            )
        except BaseException as exc:
            result["error"] = exc

    thread = threading.Thread(target=worker, daemon=True)
    thread.start()

    while thread.is_alive():
        display_gate.drain_updates(updates)
        display_gate.apply_ready()
        progress.refresh()
        time.sleep(PROGRESS_POLL_SECONDS)

    thread.join()
    display_gate.drain_updates(updates)
    while display_gate.has_pending():
        display_gate.apply_ready()
        progress.refresh()
        time.sleep(PROGRESS_POLL_SECONDS)

    if result["error"] is not None:
        if isinstance(result["error"], SystemExit):
            raise RuntimeError(str(result["error"]))
        raise result["error"]

    return result["stats"] or {}


def format_success_message(stats: dict[str, dict[str, object]]) -> str:
    if not stats:
        return f"{APP_NAME}: обробку завершено."

    parts = []
    for sheet_name, sheet_stats in stats.items():
        written = int(sheet_stats.get("written_rows", 0))
        amount = float(sheet_stats.get("total_amount", 0.0))
        parts.append(f"{sheet_name}: {written} ряд., {amount:,.2f} грн")

    return f"{APP_NAME}: колонку AN заповнено.\n" + "\n".join(parts)


def main() -> None:
    ui_root = None
    progress = None
    try:
        run_start()

        target_xlsx, start_col_idx, year, month = ask_all_inputs_groshi()
        log(f"[INPUT] TARGET file : {target_xlsx}")
        log(f"[PARAMS] year={year}, month={month}, start_col_idx={start_col_idx}")

        ui_root = Tk()
        ui_root.withdraw()
        install_frozen_executable_icon(ui_root)
        install_dark_title_bar(ui_root)
        ui_root.attributes("-topmost", True)

        progress = ProgressWindow(ui_root)
        stats = run_rewards_groshi_with_progress(
            target_xlsx=target_xlsx,
            start_col_idx=start_col_idx,
            year=year,
            month=month,
            progress=progress,
        )

        for sheet_name, sheet_stats in stats.items():
            log(f"[RESULT] {sheet_name}: {sheet_stats}")

        run_end()
        progress.show_success_then_close(
            f"{APP_NAME}: колонку AN заповнено.",
            file_name=target_xlsx.name,
        )
        progress = None

        if ui_root is not None and _widget_exists(ui_root):
            ui_root.attributes("-topmost", True)
            messagebox.showinfo("Готово!", format_success_message(stats), parent=ui_root)
    except SystemExit as exc:
        if str(exc):
            log(str(exc))
    except Exception as exc:
        error_parent = None
        if progress is not None and _widget_exists(progress.dialog):
            error_parent = progress.dialog
        elif ui_root is not None and _widget_exists(ui_root):
            error_parent = ui_root

        try:
            if error_parent is not None:
                error_parent.lift()
                error_parent.attributes("-topmost", True)
                error_parent.update()
                messagebox.showerror("Помилка", str(exc), parent=error_parent)
            else:
                root = Tk()
                root.withdraw()
                install_frozen_executable_icon(root)
                install_dark_title_bar(root)
                root.attributes("-topmost", True)
                messagebox.showerror("Помилка", str(exc), parent=root)
                root.destroy()
        except Exception:
            print(f"Помилка: {exc}")
    finally:
        if progress is not None:
            progress.close()
        if ui_root is not None and _widget_exists(ui_root):
            try:
                ui_root.destroy()
            except Exception:
                pass


if __name__ == "__main__":
    main()
