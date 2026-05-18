from typing import Callable, Dict, List, Optional, Set, Tuple, Iterable
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from tkinter import ttk, filedialog, messagebox
from datetime import date, datetime
from dataclasses import dataclass
from queue import Queue, Empty
import tkinter.font as tkfont
import sys, re, os, threading
from time import perf_counter
from pathlib import Path
import tkinter as tk


class FileAccessError(Exception):
    title = "Файл зайнятий"


_FILE_ACCESS_WINERRORS = {5, 32, 33}


def _is_file_access_error(exc: BaseException) -> bool:
    if isinstance(exc, PermissionError) or getattr(exc, "winerror", None) in _FILE_ACCESS_WINERRORS:
        return True
    text = str(exc).casefold()
    return any(
        marker in text
        for marker in (
            "permission denied",
            "access is denied",
            "used by another process",
            "file is in use",
            "locked for editing",
            "файл використовується",
            "файл зайнятий",
        )
    )


def _permission_error_path(exc: BaseException, fallback: Path | None = None) -> Path | None:
    for attr in ("filename2", "filename"):
        value = getattr(exc, attr, None)
        if value:
            try:
                return Path(value)
            except Exception:
                pass
    return Path(fallback) if fallback is not None else None


def _file_access_message(path: Path | None, action: str, exc: BaseException | None = None) -> str:
    path_text = str(path) if path else "невідомий файл"
    return (
        f"Не вдалося {action}.\n\n"
        f"Файл:\n{path_text}\n\n"
        "Причина: файл використовується іншим процесом.\n"
        "Закрийте цей файл, чи процес що його використовує..."
    )


def _raise_file_access_error(exc: BaseException, fallback: Path | None, action: str) -> None:
    path = _permission_error_path(exc, fallback)
    raise FileAccessError(_file_access_message(path, action, exc)) from exc


def ensure_file_available_for_write(path: Path, action: str = "записати файл") -> None:
    path = Path(path)
    if not path.exists():
        return
    try:
        with path.open("r+b"):
            pass
    except OSError as exc:
        if _is_file_access_error(exc):
            _raise_file_access_error(exc, path, action)
        raise


def write_text_file_with_access_check(path: str | Path, text: str, encoding: str = "utf-8-sig") -> None:
    output_path = Path(path)
    ensure_file_available_for_write(output_path, "записати файл звіту")
    try:
        with output_path.open("w", encoding=encoding) as f:
            f.write(text)
    except OSError as exc:
        if _is_file_access_error(exc):
            _raise_file_access_error(exc, output_path, "записати файл звіту")
        raise


def install_frozen_executable_icon(root, retry_ms: int = 250) -> None:
    if sys.platform != "win32" or not getattr(sys, "frozen", False):
        return

    _set_frozen_executable_icon(root)
    try:
        root.after(retry_ms, lambda: _set_frozen_executable_icon(root))
    except Exception:
        pass


def _set_frozen_executable_icon(root) -> None:
    if sys.platform != "win32" or not getattr(sys, "frozen", False):
        return

    try:
        import ctypes
        from ctypes import wintypes

        wm_seticon = 0x0080
        icon_small = 0
        icon_big = 1
        gclp_hicon = -14
        gclp_hiconsm = -34

        shell32 = ctypes.WinDLL("shell32", use_last_error=True)
        user32 = ctypes.WinDLL("user32", use_last_error=True)
        long_ptr = (
            ctypes.c_longlong if ctypes.sizeof(ctypes.c_void_p) == 8 else ctypes.c_long
        )

        shell32.ExtractIconExW.argtypes = (
            wintypes.LPCWSTR,
            ctypes.c_int,
            ctypes.POINTER(wintypes.HICON),
            ctypes.POINTER(wintypes.HICON),
            wintypes.UINT,
        )
        shell32.ExtractIconExW.restype = wintypes.UINT
        user32.SendMessageW.argtypes = (
            wintypes.HWND,
            wintypes.UINT,
            wintypes.WPARAM,
            wintypes.LPARAM,
        )
        user32.SendMessageW.restype = long_ptr
        user32.GetParent.argtypes = (wintypes.HWND,)
        user32.GetParent.restype = wintypes.HWND

        try:
            set_class_icon = user32.SetClassLongPtrW
        except AttributeError:
            set_class_icon = user32.SetClassLongW
        set_class_icon.argtypes = (wintypes.HWND, ctypes.c_int, long_ptr)
        set_class_icon.restype = long_ptr

        icon_handles = getattr(root, "_frozen_executable_icon_handles", None)
        if icon_handles is None:
            large_icons = (wintypes.HICON * 1)()
            small_icons = (wintypes.HICON * 1)()
            icon_count = shell32.ExtractIconExW(
                sys.executable,
                0,
                large_icons,
                small_icons,
                1,
            )
            if icon_count <= 0:
                return
            icon_handles = (small_icons[0], large_icons[0])
            root._frozen_executable_icon_handles = icon_handles
        small_icon, large_icon = icon_handles

        root.update_idletasks()
        hwnds = []
        for raw_hwnd in (root.winfo_id(), root.wm_frame()):
            try:
                hwnd = int(str(raw_hwnd), 0)
            except (TypeError, ValueError):
                continue
            while hwnd and hwnd not in hwnds:
                hwnds.append(hwnd)
                hwnd = user32.GetParent(hwnd)

        for hwnd in hwnds:
            if small_icon:
                user32.SendMessageW(hwnd, wm_seticon, icon_small, small_icon)
                set_class_icon(hwnd, gclp_hiconsm, small_icon)
            if large_icon:
                user32.SendMessageW(hwnd, wm_seticon, icon_big, large_icon)
                set_class_icon(hwnd, gclp_hicon, large_icon)
    except Exception:
        pass


def install_dark_title_bar(window, retry_ms: int = 80) -> None:
    if sys.platform != "win32":
        return

    def _apply() -> None:
        try:
            import ctypes

            hwnd = ctypes.windll.user32.GetParent(window.winfo_id())
            value = ctypes.c_int(1)
            for attr in (20, 19):
                ctypes.windll.dwmapi.DwmSetWindowAttribute(
                    hwnd,
                    attr,
                    ctypes.byref(value),
                    ctypes.sizeof(value),
                )
        except Exception:
            pass

    try:
        window.update_idletasks()
        _apply()
        window.after(retry_ms, _apply)
    except Exception:
        pass

# ------------------------------------------------------------
# НАЛАШТУВАННЯ / КОНФІГ
# ------------------------------------------------------------

# Тут далі можна буде спокійно додавати нові колонки і псевдоніми
COLUMN_ALIASES: Dict[str, List[str]] = {
    "serial_number": [
        "заводський номер",
        "заводський",
    ],
    "price": [
        "ціна",
    ],
    "original_cost": [
        "первісна (переоцінена) вартість",
        "первісна вартість",
        "переоцінена вартість",
    ],
}

# Скільки верхніх рядків аркуша скануємо в пошуку шапки
HEADER_SCAN_MAX_ROWS = 50

# Якщо шапка комплексна, беремо всі знайдені маркери в межах цього вікна
# і старт читання даних = найнижчий рядок маркера + 1
HEADER_GROUP_WINDOW = 5
PARAGRAPH_PREFIX = "• "
PROCESS_LOG_FILENAME = "Процес.log"
MISMATCH_FILL_COLOR = "FFFFFF00"
MISMATCH_COMMENT_LINE_RE = re.compile(r"^Файл №\d+:\s*")


def format_elapsed_seconds(elapsed_seconds: float) -> str:
    total_seconds = max(0, int(round(elapsed_seconds)))
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)

    if hours:
        return f"{hours} год {minutes:02d} хв {seconds:02d} с"
    if minutes:
        return f"{minutes} хв {seconds:02d} с"
    return f"{seconds} с"

# ------------------------------------------------------------
# ДОПОМІЖНІ СТРУКТУРИ
# ------------------------------------------------------------

@dataclass
class HeaderMatch:
    logical_name: str          # Напр. "serial_number"
    matched_alias: str         # Напр. "заводський номер"
    row: int                   # 1-based
    col: int                   # 1-based


@dataclass
class SheetParseContext:
    workbook_path: str
    sheet_name: str
    serial_col_index: int      # 1-based
    header_data_start_row: int # 1-based


@dataclass
class LoadedWorkbook:
    file_path: str
    engine: str
    workbook: object


@dataclass
class ColumnParseContext:
    workbook_path: str
    sheet_name: str
    logical_name: str
    column_index: int          # 1-based
    header_data_start_row: int # 1-based
    matched_alias: str


@dataclass
class PriceSourceContext:
    workbook_path: str
    sheet_name: str
    source_kind: str           # "price" | "original_cost"
    value_col_index: int       # 1-based
    quantity_col_index: Optional[int]
    header_data_start_row: int # 1-based
    matched_alias: str


@dataclass
class SerialPriceEvidence:
    serial: str
    row1_index: Optional[int]
    row2_index: Optional[int]
    price1_text: str
    price2_text: str
    row1_text: str
    row2_text: str
    source1_text: str
    source2_text: str
    note: str = ""


@dataclass
class DuplicateSerialEvidence:
    serial: str
    file1_rows: List[int]
    file2_rows: List[int]


# ------------------------------------------------------------
# НОРМАЛІЗАЦІЯ ТЕКСТУ
# ------------------------------------------------------------

def normalize_text(value) -> str:
    """
    Нормалізація тексту:
    - None -> ""
    - trim
    - lower
    - заміна подвійних пробілів
    """
    if value is None:
        return ""

    text = str(value).strip().lower()
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def clean_display_text(value) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        if value.time().hour == 0 and value.time().minute == 0 and value.time().second == 0 and value.time().microsecond == 0:
            return value.date().isoformat()
        return value.replace(microsecond=0).isoformat(sep=" ")

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")

    text = str(value).replace("\xa0", " ").strip()
    return re.sub(r"\s+", " ", text)


def parse_decimal(value) -> Optional[Decimal]:
    if value is None or isinstance(value, bool):
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, int):
        return Decimal(value)

    if isinstance(value, float):
        return Decimal(str(value))

    if isinstance(value, (datetime, date)):
        return None

    text = str(value).strip().replace("\xa0", "").replace(" ", "")
    if not text:
        return None

    text = text.replace("грн.", "").replace("грн", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    else:
        text = text.replace(",", ".")

    text = re.sub(r"[^0-9.\-]", "", text)
    if not text:
        return None

    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def normalize_money(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def money_to_text(value: Optional[Decimal]) -> str:
    if value is None:
        return "<немає>"
    return format(normalize_money(value), ".2f")


def is_exact_header_match(cell_value, aliases: Iterable[str]) -> Optional[str]:
    """
    Перевіряє, чи клітинка дорівнює одному з псевдонімів заголовка
    БЕЗ зайвого тексту в клітинці.
    Повертає alias, який збігся, або None.
    """
    normalized = normalize_text(cell_value)
    if not normalized:
        return None

    for alias in aliases:
        if normalized == normalize_text(alias):
            return alias
    return None


# ------------------------------------------------------------
# РОБОТА З EXCEL
# ------------------------------------------------------------

def load_workbook_generic(file_path: str) -> LoadedWorkbook:
    """
    Мінімальний універсальний loader.

    xlsx/xlsm -> openpyxl
    xls      -> xlrd

    УВАГА:
    - для .xls потрібен встановлений xlrd
    - для .xlsx/.xlsm потрібен openpyxl
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        # data_only=True -> беремо значення формул, якщо вони закешовані
        # keep_vba=True  -> для xlsm не ламаємо файл при відкритті
        try:
            wb = load_workbook(file_path, data_only=True, keep_vba=(ext == ".xlsm"))
        except OSError as exc:
            if _is_file_access_error(exc):
                _raise_file_access_error(exc, Path(file_path), "відкрити Excel-файл")
            raise
        return LoadedWorkbook(file_path=file_path, engine="openpyxl", workbook=wb)

    elif ext == ".xls":
        import xlrd
        try:
            wb = xlrd.open_workbook(file_path)
        except OSError as exc:
            if _is_file_access_error(exc):
                _raise_file_access_error(exc, Path(file_path), "відкрити Excel-файл")
            raise
        return LoadedWorkbook(file_path=file_path, engine="xlrd", workbook=wb)

    else:
        raise ValueError(f"Непідтримуваний формат файлу: {ext}")


def get_sheet_names(file_path: str) -> List[str]:
    loaded = load_workbook_generic(file_path)

    if loaded.engine == "openpyxl":
        return loaded.workbook.sheetnames

    if loaded.engine == "xlrd":
        return loaded.workbook.sheet_names()

    raise RuntimeError("Невідомий engine.")


def iter_sheet_cells(loaded: LoadedWorkbook, sheet_name: str):
    """
    Ітеруємо всі клітинки аркуша як:
    (row_index_1based, col_index_1based, cell_value)

    Це дозволяє однаково працювати і з openpyxl, і з xlrd.
    """
    if loaded.engine == "openpyxl":
        ws = loaded.workbook[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                yield cell.row, cell.column, cell.value
        return

    if loaded.engine == "xlrd":
        ws = loaded.workbook.sheet_by_name(sheet_name)
        for r in range(ws.nrows):
            for c in range(ws.ncols):
                yield r + 1, c + 1, ws.cell_value(r, c)
        return

    raise RuntimeError("Невідомий engine.")


def get_cell_value(loaded: LoadedWorkbook, sheet_name: str, row: int, col: int):
    if loaded.engine == "openpyxl":
        ws = loaded.workbook[sheet_name]
        return ws.cell(row=row, column=col).value

    if loaded.engine == "xlrd":
        ws = loaded.workbook.sheet_by_name(sheet_name)
        if row - 1 < ws.nrows and col - 1 < ws.ncols:
            return ws.cell_value(row - 1, col - 1)
        return None

    raise RuntimeError("Невідомий engine.")


def get_sheet_max_bounds(loaded: LoadedWorkbook, sheet_name: str) -> Tuple[int, int]:
    """
    Повертає (max_row, max_col)
    """
    if loaded.engine == "openpyxl":
        ws = loaded.workbook[sheet_name]
        return ws.max_row, ws.max_column

    if loaded.engine == "xlrd":
        ws = loaded.workbook.sheet_by_name(sheet_name)
        return ws.nrows, ws.ncols

    raise RuntimeError("Невідомий engine.")


# ------------------------------------------------------------
# ПОШУК ЗАГОЛОВКА + ВИЗНАЧЕННЯ СТАРТУ ЧИТАННЯ
# ------------------------------------------------------------

def find_header_matches(
    loaded: LoadedWorkbook,
    sheet_name: str,
    column_aliases: Dict[str, List[str]],
    scan_max_rows: int = HEADER_SCAN_MAX_ROWS,
) -> List[HeaderMatch]:
    """
    Шукає точні збіги заголовків у верхній частині аркуша.
    """
    matches: List[HeaderMatch] = []

    for row_idx, col_idx, cell_value in iter_sheet_cells(loaded, sheet_name):
        if row_idx > scan_max_rows:
            continue

        for logical_name, aliases in column_aliases.items():
            matched_alias = is_exact_header_match(cell_value, aliases)
            if matched_alias is not None:
                matches.append(
                    HeaderMatch(
                        logical_name=logical_name,
                        matched_alias=matched_alias,
                        row=row_idx,
                        col=col_idx,
                    )
                )

    return matches


def resolve_generic_column_context(
    loaded: LoadedWorkbook,
    sheet_name: str,
    logical_name: str,
    column_aliases: Dict[str, List[str]],
    pick: str = "first",
) -> ColumnParseContext:
    all_matches = find_header_matches(loaded, sheet_name, column_aliases)
    target_matches = [m for m in all_matches if m.logical_name == logical_name]

    if not target_matches:
        raise ValueError(
            f"Не знайдено заголовок для '{logical_name}' "
            f"на аркуші '{sheet_name}' у файлі:\n{loaded.file_path}"
        )

    target_matches.sort(key=lambda x: (x.row, x.col))
    if pick == "last":
        anchor = target_matches[-1]
    else:
        anchor = target_matches[0]
    grouped = [
        m for m in target_matches
        if abs(m.row - anchor.row) <= HEADER_GROUP_WINDOW
    ]
    bottom_header_row = max(m.row for m in grouped)
    data_start_row = bottom_header_row + 1

    return ColumnParseContext(
        workbook_path=loaded.file_path,
        sheet_name=sheet_name,
        logical_name=logical_name,
        column_index=anchor.col,
        header_data_start_row=data_start_row,
        matched_alias=anchor.matched_alias,
    )


def resolve_column_context(
    loaded: LoadedWorkbook,
    sheet_name: str,
    logical_name: str,
    column_aliases: Dict[str, List[str]],
) -> SheetParseContext:
    """
    Для заданої логічної колонки:
    - знаходимо всі точні збіги її псевдонімів
    - групуємо близькі по рядках (в межах HEADER_GROUP_WINDOW)
    - беремо найбільш правдоподібну групу
    - serial_col_index = колонка знайденого маркера
    - header_data_start_row = найнижчий рядок маркера в групі + 1

    Логіка базова, але вже придатна для дорощування.
    """
    generic_ctx = resolve_generic_column_context(
        loaded=loaded,
        sheet_name=sheet_name,
        logical_name=logical_name,
        column_aliases=column_aliases,
    )

    return SheetParseContext(
        workbook_path=generic_ctx.workbook_path,
        sheet_name=sheet_name,
        serial_col_index=generic_ctx.column_index,
        header_data_start_row=generic_ctx.header_data_start_row,
    )


# ------------------------------------------------------------
# ЧИТАННЯ СЕРІЙНИКІВ
# ------------------------------------------------------------

def split_serials(raw_value) -> List[str]:
    """
    Клітинка може містити:
    - один серійник
    - кілька серійників через кому / ; / крапку / перенос рядка / пробіли
    - уточнення в дужках, яке має лишитися з поточним серійником
    - бути пустою
    """
    text = normalize_text(raw_value)
    if not text:
        return []

    serials: List[str] = []
    buffer: List[str] = []
    paren_depth = 0
    index = 0
    text_length = len(text)

    def flush_buffer():
        value = "".join(buffer).strip()
        value = re.sub(r"^[•*\-\u2022]+\s*", "", value)
        value = re.sub(r"(?<!\s)\(", " (", value)
        if value:
            serials.append(value)
        buffer.clear()

    while index < text_length:
        char = text[index]

        if char == "(":
            paren_depth += 1
            buffer.append(char)
            index += 1
            continue

        if char == ")":
            if paren_depth > 0:
                paren_depth -= 1
            buffer.append(char)
            index += 1
            continue

        if paren_depth > 0:
            buffer.append(char)
            index += 1
            continue

        if char in ",;.\n":
            flush_buffer()
            index += 1
            continue

        if char.isspace():
            next_index = index + 1
            while next_index < text_length and text[next_index].isspace():
                next_index += 1

            if next_index < text_length and text[next_index] == "(":
                if buffer and buffer[-1] != " ":
                    buffer.append(" ")
                index = next_index
                continue

            flush_buffer()
            index = next_index
            continue

        buffer.append(char)
        index += 1

    flush_buffer()
    return serials


def normalize_serial_lookup_key(serial: str) -> str:
    value = clean_display_text(serial)
    value = re.sub(r"\s*\([^)]*\)", "", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def is_valid_serial_lookup_key(serial: str) -> bool:
    return bool(serial) and bool(re.search(r"\d", serial))


def read_serials_from_sheet(loaded: LoadedWorkbook, context: SheetParseContext) -> Set[str]:
    """
    Читає всі серійники з потрібної колонки, починаючи з data_start_row.
    """
    return set(collect_serial_occurrences(loaded, context).keys())


def collect_serial_occurrences(
    loaded: LoadedWorkbook,
    context: SheetParseContext,
) -> Dict[str, List[int]]:
    max_row, _ = get_sheet_max_bounds(loaded, context.sheet_name)
    occurrences: Dict[str, List[int]] = {}

    for row in range(context.header_data_start_row, max_row + 1):
        cell_value = get_cell_value(loaded, context.sheet_name, row, context.serial_col_index)

        row_serials = split_serials(cell_value)
        for serial in row_serials:
            lookup_key = normalize_serial_lookup_key(serial)
            if is_valid_serial_lookup_key(lookup_key):
                occurrences.setdefault(lookup_key, []).append(row)

    return occurrences


def build_duplicate_serial_evidences(
    occurrences1: Dict[str, List[int]],
    occurrences2: Dict[str, List[int]],
) -> List[DuplicateSerialEvidence]:
    duplicates: List[DuplicateSerialEvidence] = []

    for serial in sorted(set(occurrences1) | set(occurrences2)):
        file1_rows = occurrences1.get(serial, [])
        file2_rows = occurrences2.get(serial, [])
        if len(file1_rows) > 1 or len(file2_rows) > 1:
            duplicates.append(
                DuplicateSerialEvidence(
                    serial=serial,
                    file1_rows=list(file1_rows),
                    file2_rows=list(file2_rows),
                )
            )

    return duplicates


def build_serial_row_index(
    loaded: LoadedWorkbook,
    sheet_name: str,
    serial_col_index: int,
    data_start_row: int,
) -> Dict[str, int]:
    max_row, _ = get_sheet_max_bounds(loaded, sheet_name)
    row_index: Dict[str, int] = {}

    for row in range(data_start_row, max_row + 1):
        cell_value = get_cell_value(loaded, sheet_name, row, serial_col_index)
        for serial in split_serials(cell_value):
            lookup_key = normalize_serial_lookup_key(serial)
            if is_valid_serial_lookup_key(lookup_key):
                row_index.setdefault(lookup_key, row)

    return row_index


def build_row_snapshot(loaded: LoadedWorkbook, sheet_name: str, row: int) -> str:
    _max_row, max_col = get_sheet_max_bounds(loaded, sheet_name)
    values = [
        clean_display_text(get_cell_value(loaded, sheet_name, row, col))
        for col in range(1, max_col + 1)
    ]

    while values and not values[-1]:
        values.pop()

    return ", ".join(values)


def resolve_price_source_context(
    loaded: LoadedWorkbook,
    sheet_name: str,
) -> PriceSourceContext:
    direct_ctx: Optional[ColumnParseContext] = None
    cost_ctx: Optional[ColumnParseContext] = None

    try:
        direct_ctx = resolve_generic_column_context(
            loaded=loaded,
            sheet_name=sheet_name,
            logical_name="price",
            column_aliases=COLUMN_ALIASES,
        )
    except ValueError:
        direct_ctx = None

    try:
        cost_ctx = resolve_generic_column_context(
            loaded=loaded,
            sheet_name=sheet_name,
            logical_name="original_cost",
            column_aliases=COLUMN_ALIASES,
            pick="last",
        )
    except ValueError:
        cost_ctx = None

    if direct_ctx is not None:
        return PriceSourceContext(
            workbook_path=loaded.file_path,
            sheet_name=sheet_name,
            source_kind="price",
            value_col_index=direct_ctx.column_index,
            quantity_col_index=None,
            header_data_start_row=direct_ctx.header_data_start_row,
            matched_alias=direct_ctx.matched_alias,
        )

    if cost_ctx is not None:
        quantity_col_index = cost_ctx.column_index - 1
        if quantity_col_index < 1:
            raise ValueError(
                f"Для колонки '{cost_ctx.matched_alias}' на аркуші '{sheet_name}' "
                f"у файлі '{loaded.file_path}' немає сусідньої колонки зліва для кількості."
            )
        return PriceSourceContext(
            workbook_path=loaded.file_path,
            sheet_name=sheet_name,
            source_kind="original_cost",
            value_col_index=cost_ctx.column_index,
            quantity_col_index=quantity_col_index,
            header_data_start_row=cost_ctx.header_data_start_row,
            matched_alias=cost_ctx.matched_alias,
        )

    raise ValueError(
        f"Не знайдено ні прямий заголовок 'ціна', ні "
        f"'первісна (переоцінена) вартість' на аркуші '{sheet_name}' "
        f"у файлі:\n{loaded.file_path}"
    )


def extract_price_from_row(
    loaded: LoadedWorkbook,
    sheet_name: str,
    row: int,
    price_context: PriceSourceContext,
) -> Decimal:
    raw_value = get_cell_value(loaded, sheet_name, row, price_context.value_col_index)
    numeric_value = parse_decimal(raw_value)
    if numeric_value is None:
        raise ValueError(
            f"Не вдалося прочитати значення ціни у рядку {row} "
            f"(колонка {price_context.value_col_index})."
        )

    if price_context.source_kind == "price":
        return normalize_money(numeric_value)

    if price_context.quantity_col_index is None:
        raise ValueError("Не задано колонку кількості для обчислення ціни.")

    raw_quantity = get_cell_value(loaded, sheet_name, row, price_context.quantity_col_index)
    quantity_value = parse_decimal(raw_quantity)
    if quantity_value is None:
        raise ValueError(
            f"Не вдалося прочитати кількість у рядку {row} "
            f"(колонка {price_context.quantity_col_index})."
        )
    if quantity_value == 0:
        raise ValueError(
            f"Кількість у рядку {row} дорівнює 0, обчислити ціну неможливо."
        )

    return normalize_money(numeric_value / quantity_value)


def build_price_source_label(context: PriceSourceContext) -> str:
    if context.source_kind == "price":
        return f"пряма ціна ('{context.matched_alias}')"
    quantity_label = f"колонка {context.quantity_col_index}"
    return (
        f"первісна/переоцінена вартість ('{context.matched_alias}') / "
        f"кількість зліва ({quantity_label})"
    )


def compare_prices_for_common_serials(
    loaded1: LoadedWorkbook,
    loaded2: LoadedWorkbook,
    sheet1: str,
    sheet2: str,
    serial_ctx1: SheetParseContext,
    serial_ctx2: SheetParseContext,
    price_ctx1: PriceSourceContext,
    price_ctx2: PriceSourceContext,
    serials: List[str],
    progress_callback: Optional[Callable[[int, int, str], None]] = None,
) -> Tuple[List[SerialPriceEvidence], List[SerialPriceEvidence]]:
    data_start_row1 = max(serial_ctx1.header_data_start_row, price_ctx1.header_data_start_row)
    data_start_row2 = max(serial_ctx2.header_data_start_row, price_ctx2.header_data_start_row)

    row_index1 = build_serial_row_index(loaded1, sheet1, serial_ctx1.serial_col_index, data_start_row1)
    row_index2 = build_serial_row_index(loaded2, sheet2, serial_ctx2.serial_col_index, data_start_row2)

    matched: List[SerialPriceEvidence] = []
    mismatched: List[SerialPriceEvidence] = []

    total_serials = len(serials)
    source1_label = build_price_source_label(price_ctx1)
    source2_label = build_price_source_label(price_ctx2)

    for idx, serial in enumerate(serials, start=1):
        if progress_callback is not None:
            progress_callback(idx, total_serials, serial)

        row1 = row_index1.get(serial)
        row2 = row_index2.get(serial)

        row1_text = build_row_snapshot(loaded1, sheet1, row1) if row1 is not None else ""
        row2_text = build_row_snapshot(loaded2, sheet2, row2) if row2 is not None else ""
        note_parts: List[str] = []
        price1: Optional[Decimal] = None
        price2: Optional[Decimal] = None

        if row1 is None:
            note_parts.append("серійник не знайдено повторно у Файлі №1")
        else:
            try:
                price1 = extract_price_from_row(loaded1, sheet1, row1, price_ctx1)
            except Exception as exc:
                note_parts.append(f"Файл №1: {exc}")

        if row2 is None:
            note_parts.append("серійник не знайдено повторно у Файлі №2")
        else:
            try:
                price2 = extract_price_from_row(loaded2, sheet2, row2, price_ctx2)
            except Exception as exc:
                note_parts.append(f"Файл №2: {exc}")

        evidence = SerialPriceEvidence(
            serial=serial,
            row1_index=row1,
            row2_index=row2,
            price1_text=money_to_text(price1),
            price2_text=money_to_text(price2),
            row1_text=row1_text,
            row2_text=row2_text,
            source1_text=source1_label,
            source2_text=source2_label,
            note="; ".join(note_parts),
        )

        if price1 is not None and price2 is not None and normalize_money(price1) == normalize_money(price2):
            matched.append(evidence)
        else:
            mismatched.append(evidence)

    return matched, mismatched


def build_mismatch_row_links(
    evidences: List[SerialPriceEvidence],
) -> Tuple[Dict[int, Set[int]], Dict[int, Set[int]]]:
    row_links1: Dict[int, Set[int]] = {}
    row_links2: Dict[int, Set[int]] = {}

    for entry in evidences:
        if entry.row1_index is not None:
            row_links1.setdefault(entry.row1_index, set())
        if entry.row2_index is not None:
            row_links2.setdefault(entry.row2_index, set())
        if entry.row1_index is not None and entry.row2_index is not None:
            row_links1[entry.row1_index].add(entry.row2_index)
            row_links2[entry.row2_index].add(entry.row1_index)

    return row_links1, row_links2


def build_mismatch_comment(other_file_number: int, other_rows: Set[int]) -> str:
    sorted_rows = sorted(other_rows)
    if not sorted_rows:
        return f"Файл №{other_file_number}: не знайдено"

    rows_text = ", ".join(str(row) for row in sorted_rows)
    return f"Файл №{other_file_number}: {rows_text}"


def merge_mismatch_comment_text(existing_text: str, mismatch_line: str) -> str:
    preserved_lines: List[str] = []

    for line in (existing_text or "").splitlines():
        if not MISMATCH_COMMENT_LINE_RE.match(line.strip()):
            preserved_lines.append(line.rstrip())

    while preserved_lines and not preserved_lines[-1].strip():
        preserved_lines.pop()

    if preserved_lines:
        preserved_lines.append("")
    preserved_lines.append(mismatch_line)
    return "\n".join(preserved_lines)


def annotate_mismatched_rows_in_workbook(
    file_number: int,
    file_path: str,
    sheet_name: str,
    serial_col_index: int,
    row_links: Dict[int, Set[int]],
    other_file_number: int,
) -> Optional[str]:
    if not row_links:
        return None

    source_caption = build_log_source_caption(file_number, file_path, sheet_name)
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in (".xlsx", ".xlsm"):
        return (
            f"{source_caption}: автоматичне підсвічування та примітки пропущено, "
            f"бо формат {ext} не підтримує запис."
        )

    try:
        from openpyxl import load_workbook
        from openpyxl.comments import Comment
        from openpyxl.styles import PatternFill
    except Exception as exc:
        raise ValueError(
            f"Не вдалося підготувати редагування {source_caption}: недоступний openpyxl ({exc})."
        ) from exc

    workbook = None
    try:
        ensure_file_available_for_write(Path(file_path), "зберегти позначки в Excel-файлі")
        workbook = load_workbook(
            file_path,
            data_only=False,
            keep_vba=(ext == ".xlsm"),
            keep_links=True,
        )
        worksheet = workbook[sheet_name]
        fill = PatternFill(fill_type="solid", fgColor=MISMATCH_FILL_COLOR)
        max_col = max(worksheet.max_column, serial_col_index)

        for row_index, other_rows in sorted(row_links.items()):
            for col_index in range(1, max_col + 1):
                worksheet.cell(row=row_index, column=col_index).fill = fill

            serial_cell = worksheet.cell(row=row_index, column=serial_col_index)
            existing_comment = serial_cell.comment.text if serial_cell.comment is not None else ""
            serial_cell.comment = Comment(
                merge_mismatch_comment_text(
                    existing_comment,
                    build_mismatch_comment(other_file_number, other_rows),
                ),
                "ECOmpare",
            )

        try:
            workbook.save(file_path)
        except OSError as exc:
            if _is_file_access_error(exc):
                _raise_file_access_error(exc, Path(file_path), "зберегти позначки в Excel-файлі")
            raise
    except FileAccessError:
        raise
    except KeyError as exc:
        raise ValueError(
            f"Не знайдено аркуш {quote_log_value(sheet_name)} у {source_caption} під час запису позначок."
        ) from exc
    except Exception as exc:
        raise ValueError(f"Не вдалося оновити {source_caption}: {exc}") from exc
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass

    row_word = "рядок" if len(row_links) == 1 else "рядки"
    return f"{source_caption}: позначено {len(row_links)} {row_word} жовтою заливкою та примітками."


def annotate_mismatched_price_rows(
    file1_path: str,
    file2_path: str,
    sheet1_name: str,
    sheet2_name: str,
    serial_ctx1: SheetParseContext,
    serial_ctx2: SheetParseContext,
    evidences: List[SerialPriceEvidence],
) -> List[str]:
    row_links1, row_links2 = build_mismatch_row_links(evidences)
    messages: List[str] = []

    for message in (
        annotate_mismatched_rows_in_workbook(
            file_number=1,
            file_path=file1_path,
            sheet_name=sheet1_name,
            serial_col_index=serial_ctx1.serial_col_index,
            row_links=row_links1,
            other_file_number=2,
        ),
        annotate_mismatched_rows_in_workbook(
            file_number=2,
            file_path=file2_path,
            sheet_name=sheet2_name,
            serial_col_index=serial_ctx2.serial_col_index,
            row_links=row_links2,
            other_file_number=1,
        ),
    ):
        if message:
            messages.append(message)

    return messages


# ------------------------------------------------------------
# ЗВІТ
# ------------------------------------------------------------

def build_serials_file_text(header: str, serials: List[str]) -> str:
    lines = [header, ""]
    if serials:
        lines.extend(f"{PARAGRAPH_PREFIX}{serial}" for serial in serials)
    else:
        lines.append("<немає>")
    return "\n".join(lines) + "\n"


def build_price_report_text(
    header: str,
    evidences: List[SerialPriceEvidence],
) -> str:
    lines = [header, ""]

    if not evidences:
        lines.append("<немає>")
        return "\n".join(lines) + "\n"

    for entry in evidences:
        lines.append(f"Серійник: {entry.serial}")
        lines.append(
            f"{PARAGRAPH_PREFIX}Файл №1 | ціна: {entry.price1_text} | рядок: {entry.row1_index if entry.row1_index is not None else '<не знайдено>'} | джерело: {entry.source1_text}"
        )
        lines.append(entry.row1_text or "<порожній рядок>")
        lines.append(
            f"{PARAGRAPH_PREFIX}Файл №2 | ціна: {entry.price2_text} | рядок: {entry.row2_index if entry.row2_index is not None else '<не знайдено>'} | джерело: {entry.source2_text}"
        )
        lines.append(entry.row2_text or "<порожній рядок>")
        if entry.note:
            lines.append(f"Примітка: {entry.note}")
        lines.append("")

    return "\n".join(lines).rstrip() + "\n"


def build_duplicate_report_text(
    header: str,
    evidences: List[DuplicateSerialEvidence],
) -> str:
    lines = [header, ""]

    if not evidences:
        lines.append("<немає>")
        return "\n".join(lines) + "\n"

    for entry in evidences:
        file1_rows = ", ".join(str(row) for row in entry.file1_rows) if entry.file1_rows else "<немає>"
        file2_rows = ", ".join(str(row) for row in entry.file2_rows) if entry.file2_rows else "<немає>"
        lines.append(f"Серійник: {entry.serial}")
        lines.append(f"{PARAGRAPH_PREFIX}Файл №1 | входжень: {len(entry.file1_rows)} | рядки: {file1_rows}")
        lines.append(f"{PARAGRAPH_PREFIX}Файл №2 | входжень: {len(entry.file2_rows)} | рядки: {file2_rows}")
        lines.append("")

    return "\n".join(lines).rstrip() + "\n"


def quote_log_value(value: str) -> str:
    text = str(value or "").replace('"', '\\"')
    return f'"{text}"'


def paths_refer_same_workbook(file1_path: str, file2_path: str) -> bool:
    if not file1_path or not file2_path:
        return False
    return os.path.normcase(os.path.abspath(file1_path)) == os.path.normcase(os.path.abspath(file2_path))


def build_log_source_caption(
    file_number: int,
    file_path: str,
    sheet_name: str,
    grammatical_case: str = "nominative",
    include_file_name: bool = True,
) -> str:
    sheet_label = sheet_name or "<не вказано>"
    file_word = "Файл" if grammatical_case == "nominative" else "файлі"
    if not include_file_name:
        return f"{file_word} №{file_number} (аркуш: {quote_log_value(sheet_label)})"

    file_name = os.path.basename(file_path) or file_path
    return (
        f"{file_word} №{file_number} "
        f"({quote_log_value(file_name)}, аркуш: {quote_log_value(sheet_label)})"
    )


def save_reports_near_file2(
    file1_path: str,
    file2_path: str,
    sheet1_name: str,
    sheet2_name: str,
    duplicate_serials: List[DuplicateSerialEvidence],
    only_in_file1: List[str],
    only_in_file2: List[str],
    in_both: List[str],
) -> Tuple[str, str, str, str]:
    out_dir = os.path.dirname(file2_path)
    include_file_name = not paths_refer_same_workbook(file1_path, file2_path)
    source1 = build_log_source_caption(
        1, file1_path, sheet1_name, grammatical_case="locative", include_file_name=include_file_name
    )
    source2 = build_log_source_caption(
        2, file2_path, sheet2_name, grammatical_case="locative", include_file_name=include_file_name
    )
    out_path0 = os.path.join(out_dir, "Серійники №0.log")
    out_path1 = os.path.join(out_dir, "Серійники №1.log")
    out_path2 = os.path.join(out_dir, "Серійники №2.log")
    out_path3 = os.path.join(out_dir, "Серійники №3.log")

    write_text_file_with_access_check(
        out_path0,
        build_duplicate_report_text(
            f"--- ДУБЛІКАТИ СЕРІЙНИКІВ у {source1} та у {source2}:",
            duplicate_serials,
        )
    )

    write_text_file_with_access_check(
        out_path1,
        build_serials_file_text(
            f"--- СЕРІЙНИКИ, що існують у {source1}, але немає у {source2}:",
            only_in_file1,
        )
    )

    write_text_file_with_access_check(
        out_path2,
        build_serials_file_text(
            f"--- СЕРІЙНИКИ, що існують у {source2}, але немає у {source1}:",
            only_in_file2,
        )
    )

    write_text_file_with_access_check(
        out_path3,
        build_serials_file_text(
            f"--- СЕРІЙНИКИ, що існують і у {source1}, і у {source2}:",
            in_both,
        )
    )

    return out_path0, out_path1, out_path2, out_path3


def save_price_reports_near_file2(
    file1_path: str,
    file2_path: str,
    sheet1_name: str,
    sheet2_name: str,
    matched: List[SerialPriceEvidence],
    mismatched: List[SerialPriceEvidence],
) -> Tuple[str, str]:
    out_dir = os.path.dirname(file2_path)
    include_file_name = not paths_refer_same_workbook(file1_path, file2_path)
    source1 = build_log_source_caption(
        1, file1_path, sheet1_name, grammatical_case="locative", include_file_name=include_file_name
    )
    source2 = build_log_source_caption(
        2, file2_path, sheet2_name, grammatical_case="locative", include_file_name=include_file_name
    )
    out_path4 = os.path.join(out_dir, "Серійники №4.log")
    out_path5 = os.path.join(out_dir, "Серійники №5.log")

    write_text_file_with_access_check(
        out_path4,
        build_price_report_text(
            f"--- СЕРІЙНИКИ, де ціни збігаються у {source1} та у {source2}:",
            matched,
        )
    )

    write_text_file_with_access_check(
        out_path5,
        build_price_report_text(
            f"--- СЕРІЙНИКИ, де ціни НЕ збігаються у {source1} та у {source2}:",
            mismatched,
        )
    )

    return out_path4, out_path5


# ------------------------------------------------------------
# GUI / DIALOGS
# ------------------------------------------------------------

# Тут можна швидко підкрутити геометрію стартового вікна.
FILE_PATH_ENTRY_WIDTH = 40
ACTION_BUTTON_WIDTH = 11
SHEET_BUTTON_WIDTH = 5
LAUNCH_WINDOW_MIN_WIDTH = 500
LAUNCH_WINDOW_MIN_HEIGHT = 290
PROGRESS_WINDOW_WIDTH = 640
PROGRESS_WINDOW_HEIGHT = 236
HEADER_PAD_X = 20
HEADER_PAD_Y = 10
BODY_PADDING = (16, 14, 16, 16)
PANEL_PADDING = (14, 12, 14, 12)
ROW_PAD_X = 8
ROW_PAD_Y = 5


def configure_app_style(root: tk.Misc) -> Dict[str, str]:
    colors = {
        "window": "#F4F7FB",
        "panel": "#FFFFFF",
        "border": "#D6DEE8",
        "text": "#15202B",
        "muted": "#617084",
        "accent": "#0F766E",
        "accent_dark": "#0B5D56",
        "header": "#103C3A",
        "header_badge": "#0F766E",
    }

    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except Exception:
        pass

    try:
        style.configure("LaunchRoot.TFrame", background=colors["window"])
        style.configure("LaunchPanel.TFrame", background=colors["panel"])
        style.configure("LaunchSection.TLabel", background=colors["panel"], foreground=colors["text"], font=("Segoe UI", 12, "bold"))
        style.configure("LaunchField.TLabel", background=colors["panel"], foreground=colors["text"], font=("Segoe UI", 10, "bold"))
        style.configure("LaunchMuted.TLabel", background=colors["panel"], foreground=colors["muted"], font=("Segoe UI", 9))
        style.configure("Header.TLabel", background=colors["panel"], foreground=colors["text"], font=("Segoe UI", 12, "bold"))
        style.configure("Body.TLabel", background=colors["panel"], foreground=colors["text"], font=("Segoe UI", 10))
        style.configure("ProgressHeader.TLabel", background=colors["panel"], foreground=colors["text"], font=("Segoe UI", 12, "bold"))
        style.configure("ProgressBody.TLabel", background=colors["panel"], foreground=colors["text"], font=("Segoe UI", 10))
        style.configure("ProgressFile.TLabel", background=colors["panel"], foreground=colors["muted"], font=("Segoe UI", 9))
        style.configure("ProgressCount.TLabel", background=colors["panel"], foreground=colors["accent_dark"], font=("Segoe UI", 10, "bold"))
        style.configure(
            "Launch.Horizontal.TProgressbar",
            background=colors["header_badge"],
            troughcolor="#E8EEF5",
            bordercolor=colors["border"],
            lightcolor=colors["header_badge"],
            darkcolor=colors["header_badge"],
            thickness=12,
        )
        style.configure(
            "Launch.TEntry",
            fieldbackground="#FFFFFF",
            foreground=colors["text"],
            bordercolor=colors["border"],
            lightcolor=colors["border"],
            darkcolor=colors["border"],
            padding=(8, 6),
        )
        style.map(
            "Launch.TEntry",
            fieldbackground=[("readonly", "#FFFFFF")],
            foreground=[("readonly", colors["text"])],
        )
        style.configure(
            "Action.TButton",
            font=("Segoe UI", 10, "bold"),
            padding=(14, 8),
            foreground="#FFFFFF",
            background="#053D39",
            bordercolor=colors["accent_dark"],
            lightcolor=colors["accent_dark"],
            darkcolor=colors["accent_dark"],
            focuscolor=colors["accent_dark"],
        )
        style.map(
            "Action.TButton",
            background=[("pressed", colors["accent_dark"]), ("active", colors["accent_dark"])],
            foreground=[("pressed", "#FFFFFF"), ("active", "#FFFFFF")],
        )
        style.configure(
            "Launch.TButton",
            font=("Segoe UI", 10, "bold"),
            padding=(16, 10),
            foreground="#FFFFFF",
            background="#053D39",
            bordercolor=colors["accent_dark"],
            lightcolor=colors["accent_dark"],
            darkcolor=colors["accent_dark"],
            focuscolor=colors["accent_dark"],
        )
        style.map(
            "Launch.TButton",
            background=[("pressed", colors["accent_dark"]), ("active", colors["accent_dark"])],
            foreground=[("pressed", "#FFFFFF"), ("active", "#FFFFFF")],
        )
    except Exception:
        pass

    return colors


def widget_exists(widget: Optional[tk.Misc]) -> bool:
    if widget is None:
        return False
    try:
        return bool(widget.winfo_exists())
    except (tk.TclError, RuntimeError):
        return False


def destroy_widget(widget: Optional[tk.Misc]) -> None:
    if not widget_exists(widget):
        return
    try:
        widget.destroy()
    except (tk.TclError, RuntimeError):
        pass


def prepare_dialog_parent(root: tk.Misc) -> None:
    if not widget_exists(root):
        return
    try:
        try:
            current_grab = root.grab_current()
            if current_grab is not None and current_grab is not root:
                current_grab.grab_release()
        except (tk.TclError, RuntimeError):
            pass
        if hasattr(root, "deiconify") and not root.winfo_ismapped():
            root.deiconify()
        root.lift()
        root.attributes("-topmost", True)
        root.update()
    except (tk.TclError, RuntimeError):
        pass


def center_window(
    window: tk.Misc,
    parent: Optional[tk.Misc] = None,
    width: Optional[int] = None,
    height: Optional[int] = None,
) -> None:
    if not widget_exists(window):
        return
    window.update_idletasks()

    if width is None:
        width = window.winfo_width()
        if width <= 1:
            width = window.winfo_reqwidth()

    if height is None:
        height = window.winfo_height()
        if height <= 1:
            height = window.winfo_reqheight()

    width = max(width, window.winfo_reqwidth())
    height = max(height, window.winfo_reqheight())
    window.geometry(f"{width}x{height}")


def show_topmost_message(
    kind: str,
    title: str,
    text: str,
    parent: Optional[tk.Misc] = None,
    reveal_parent: bool = True,
) -> None:
    if widget_exists(parent):
        if reveal_parent:
            prepare_dialog_parent(parent)
        else:
            try:
                current_grab = parent.grab_current()
                if current_grab is not None and current_grab is not parent:
                    current_grab.grab_release()
            except (tk.TclError, RuntimeError):
                pass
            try:
                parent.attributes("-topmost", True)
                parent.update_idletasks()
            except (tk.TclError, RuntimeError):
                pass
        if kind == "info":
            messagebox.showinfo(title, text, parent=parent)
        elif kind == "warning":
            messagebox.showwarning(title, text, parent=parent)
        else:
            messagebox.showerror(title, text, parent=parent)
        return

    root = tk.Tk()
    root.withdraw()
    install_frozen_executable_icon(root)
    install_dark_title_bar(root)
    root.attributes("-topmost", True)
    try:
        if kind == "info":
            messagebox.showinfo(title, text, parent=root)
        elif kind == "warning":
            messagebox.showwarning(title, text, parent=root)
        else:
            messagebox.showerror(title, text, parent=root)
    finally:
        destroy_widget(root)


class ProgressWindow:
    def __init__(self, owner: tk.Misc):
        self.dialog = tk.Toplevel(owner)
        self.dialog.withdraw()
        install_frozen_executable_icon(self.dialog)
        self.dialog.title("ECOmpare - Обробка")
        self.dialog.resizable(False, False)
        self.dialog.protocol("WM_DELETE_WINDOW", self.close)
        colors = configure_app_style(self.dialog)
        self.dialog.configure(bg=colors["window"])
        install_dark_title_bar(self.dialog)

        container = ttk.Frame(self.dialog, style="LaunchRoot.TFrame")
        container.pack(fill="both", expand=True)

        header = tk.Frame(container, bg=colors["header"], padx=HEADER_PAD_X, pady=HEADER_PAD_Y)
        header.pack(fill="x")
        header.grid_columnconfigure(0, weight=1)
        tk.Label(
            header,
            text="ECOmpare",
            bg=colors["header"],
            fg="#FFFFFF",
            font=("Segoe UI Semibold", 15),
            anchor="w",
        ).grid(row=0, column=0, sticky="w")
        tk.Label(
            header,
            text="Порівняння Excel виконується. Будь ласка, дочекайтеся завершення.",
            bg=colors["header"],
            fg="#D7FBF5",
            font=("Segoe UI", 9),
            anchor="w",
        ).grid(row=1, column=0, sticky="w", pady=(3, 4))
        badge_shell = tk.Frame(header, bg=colors["accent_dark"], padx=1, pady=1)
        badge = tk.Label(
            badge_shell,
            text="ОБРОБКА",
            bg="#053D39",
            fg="#FFFFFF",
            font=("Segoe UI", 11, "bold"),
            padx=9,
            pady=4,
        )
        badge.pack()
        badge_shell.place(relx=1.0, x=0, y=3, anchor="ne")

        body = ttk.Frame(container, style="LaunchRoot.TFrame", padding=BODY_PADDING)
        body.pack(fill="both", expand=True)
        panel_shell = tk.Frame(body, bg=colors["border"], padx=1, pady=1)
        panel_shell.pack(fill="both", expand=True)
        frame = ttk.Frame(panel_shell, style="LaunchPanel.TFrame", padding=PANEL_PADDING)
        frame.pack(fill="both", expand=True)
        frame.grid_columnconfigure(0, weight=1)

        self.header_var = tk.StringVar(value="Підготовка...")
        self.detail_var = tk.StringVar(value="Будь ласка, зачекайте")
        self.file_var = tk.StringVar(value="")
        self.count_var = tk.StringVar(value="0 / 0")

        ttk.Label(frame, textvariable=self.header_var, style="ProgressHeader.TLabel").grid(
            row=0, column=0, padx=(0, 12), sticky="w"
        )
        ttk.Label(frame, textvariable=self.count_var, style="ProgressCount.TLabel").grid(
            row=0, column=1, sticky="e"
        )
        ttk.Label(frame, textvariable=self.detail_var, style="ProgressBody.TLabel", wraplength=552).grid(
            row=1, column=0, columnspan=2, pady=(8, 6), sticky="w"
        )
        ttk.Label(frame, textvariable=self.file_var, style="ProgressFile.TLabel", wraplength=552).grid(
            row=2, column=0, columnspan=2, pady=(0, 12), sticky="w"
        )

        self.progress = ttk.Progressbar(
            frame,
            orient="horizontal",
            mode="determinate",
            maximum=100,
            style="Launch.Horizontal.TProgressbar",
        )
        self.progress.grid(row=3, column=0, columnspan=2, sticky="we")
        center_window(
            self.dialog,
            parent=owner,
            width=PROGRESS_WINDOW_WIDTH,
            height=PROGRESS_WINDOW_HEIGHT,
        )
        self._show_in_front()
        self.refresh()

    def _show_in_front(self):
        try:
            self.dialog.deiconify()
            self.dialog.lift()
            self.dialog.attributes("-topmost", True)
            self.dialog.update_idletasks()
            self.dialog.attributes("-topmost", False)
            self.dialog.focus_force()
        except Exception:
            pass

    def update(
        self,
        header: Optional[str] = None,
        detail: Optional[str] = None,
        file_name: Optional[str] = None,
        current: Optional[int] = None,
        total: Optional[int] = None,
    ):
        try:
            if not self.dialog.winfo_exists():
                return
            if header is not None:
                self.header_var.set(header)
            if detail is not None:
                self.detail_var.set(detail)
            if file_name is not None:
                self.file_var.set(file_name)
            if current is not None and total is not None:
                safe_total = total if total > 0 else 1
                pct = max(0.0, min(100.0, (current / safe_total) * 100.0))
                self.progress["value"] = pct
                self.count_var.set(f"{current} / {total}")
            self.refresh()
        except tk.TclError:
            return

    def refresh(self):
        try:
            self.dialog.update_idletasks()
        except Exception:
            pass

    def close(self):
        try:
            self.dialog.destroy()
        except Exception:
            pass


class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("ПАРАМЕТРИ: ECOmpare")
        self.root.resizable(False, False)
        self.root.attributes("-topmost", True)
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        self.colors = configure_app_style(self.root)
        self.root.configure(bg=self.colors["window"])
        install_dark_title_bar(self.root)

        self.file1_var = tk.StringVar()
        self.file2_var = tk.StringVar()
        self.file1_display_var = tk.StringVar(value="")
        self.file2_display_var = tk.StringVar(value="")
        self.sheet1_var = tk.StringVar()
        self.sheet2_var = tk.StringVar()
        self.file1_sheet_names: List[str] = []
        self.file2_sheet_names: List[str] = []
        self.sheet_names_cache: Dict[str, List[str]] = {}
        self.sheet_menu_cache: Dict[Tuple[str, str], tk.Menu] = {}
        self.compare_queue: Queue = Queue()
        self.compare_in_progress = False
        self.progress_window: Optional[ProgressWindow] = None
        self.process_log_path: Optional[str] = None
        self.process_log_lock = threading.Lock()
        self.stage_started_at: Dict[str, float] = {}
        self.run_started_at: Optional[float] = None

        self._build_ui()
        self.root.grab_set()
        self.root.focus_force()

    def on_close(self):
        if self.progress_window is not None:
            self.progress_window.close()
            self.progress_window = None

        self._invalidate_sheet_menu_cache("file1")
        self._invalidate_sheet_menu_cache("file2")

        try:
            self.root.grab_release()
        except Exception:
            pass

        try:
            self.root.withdraw()
        except Exception:
            pass

        try:
            self.root.quit()
        except Exception:
            pass

        destroy_widget(self.root)

    def _append_process_log(self, message: str) -> None:
        if not self.process_log_path:
            return

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{timestamp}] {message}\n"

        with self.process_log_lock:
            with open(self.process_log_path, "a", encoding="utf-8") as f:
                f.write(line)

    def _initialize_process_log(self, file1: str, file2: str, sheet1: str, sheet2: str) -> None:
        out_dir = os.path.dirname(file2) or os.getcwd()
        self.process_log_path = os.path.join(out_dir, PROCESS_LOG_FILENAME)

        separator = "=" * 72
        with self.process_log_lock:
            with open(self.process_log_path, "a", encoding="utf-8") as f:
                f.write(f"{separator}\n")

        self._append_process_log("START RUN")
        self._append_process_log(build_log_source_caption(1, file1, sheet1))
        self._append_process_log(build_log_source_caption(2, file2, sheet2))

    def _begin_run_timing(self) -> None:
        self.run_started_at = perf_counter()
        self.stage_started_at.clear()

    def _begin_stage(self, stage_key: str, message: str) -> None:
        self.stage_started_at[stage_key] = perf_counter()
        self._append_process_log(message)

    def _finish_stage(self, stage_key: str, message: str, outcome: str = "завершено") -> None:
        started_at = self.stage_started_at.pop(stage_key, None)
        if started_at is None:
            return
        elapsed = perf_counter() - started_at
        self._append_process_log(f"{message} {outcome} за {format_elapsed_seconds(elapsed)}")

    def _finish_run_timing(self) -> None:
        if self.run_started_at is None:
            return
        elapsed = perf_counter() - self.run_started_at
        self.run_started_at = None
        self.stage_started_at.clear()
        self._append_process_log(f"RUN завершено за {format_elapsed_seconds(elapsed)}")

    def _build_ui(self):
        colors = self.colors

        def bind_sheet_button_hover(button: tk.Button) -> None:
            def on_enter(_event=None) -> None:
                if str(button.cget("state")) != "disabled":
                    button.configure(bg="#D7E0EA")

            def on_leave(_event=None) -> None:
                button.configure(bg="#FFFFFF")

            button.bind("<Enter>", on_enter, add="+")
            button.bind("<Leave>", on_leave, add="+")

        def build_file_field(
            parent: tk.Misc,
            row: int,
            display_var: tk.StringVar,
            sheet_command,
        ) -> Tuple[tk.Entry, tk.Button]:
            field_shell = tk.Frame(parent, bg=colors["border"], bd=0, padx=1, pady=1)
            field_shell.grid(row=row, column=1, padx=(0, ROW_PAD_X), pady=ROW_PAD_Y, sticky="we")
            field_shell.grid_columnconfigure(0, weight=1)

            field_panel = tk.Frame(field_shell, bg="#FFFFFF", bd=0)
            field_panel.grid(row=0, column=0, sticky="we")
            field_panel.grid_columnconfigure(0, weight=1)

            entry = tk.Entry(
                field_panel,
                bg="#FFFFFF",
                bd=0,
                fg=colors["text"],
                font=("Segoe UI", 10),
                highlightthickness=0,
                insertbackground=colors["text"],
                readonlybackground="#FFFFFF",
                relief="flat",
                state="readonly",
                textvariable=display_var,
                width=FILE_PATH_ENTRY_WIDTH,
            )
            entry.grid(row=0, column=0, padx=(8, 6), pady=6, sticky="we")

            separator = tk.Frame(field_panel, bg=colors["border"], width=1)
            separator.grid(row=0, column=1, sticky="ns")

            sheet_button = tk.Button(
                field_panel,
                activebackground="#FFFFFF",
                activeforeground=colors["text"],
                bg="#FFFFFF",
                bd=0,
                command=sheet_command,
                cursor="arrow",
                disabledforeground=colors["muted"],
                fg=colors["muted"],
                font=("Segoe UI", 9, "bold"),
                highlightthickness=0,
                padx=5,
                pady=4,
                relief="flat",
                state="disabled",
                text="Ар.",
                width=SHEET_BUTTON_WIDTH,
            )
            sheet_button.grid(row=0, column=2, sticky="ns")
            bind_sheet_button_hover(sheet_button)
            return entry, sheet_button

        surface = ttk.Frame(self.root, style="LaunchRoot.TFrame")
        surface.grid(row=0, column=0, sticky="nsew")
        surface.grid_columnconfigure(0, weight=1)

        header = tk.Frame(surface, bg=colors["header"], padx=HEADER_PAD_X, pady=HEADER_PAD_Y)
        header.grid(row=0, column=0, sticky="we")
        header.grid_columnconfigure(0, weight=1)
        tk.Label(
            header,
            text="ECOmpare",
            bg=colors["header"],
            fg="#FFFFFF",
            font=("Segoe UI Semibold", 15),
            anchor="w",
        ).grid(row=0, column=0, sticky="w")
        tk.Label(
            header,
            text="Підготовка порівняння даних Excel.",
            bg=colors["header"],
            fg="#D7FBF5",
            font=("Segoe UI", 9),
            anchor="w",
        ).grid(row=1, column=0, sticky="w", pady=(3, 4))
        badge_shell = tk.Frame(header, bg=colors["accent_dark"], padx=1, pady=1)
        badge = tk.Label(
            badge_shell,
            text="ПАРАМЕТРИ",
            bg="#053D39",
            fg="#FFFFFF",
            font=("Segoe UI", 11, "bold"),
            padx=9,
            pady=4,
        )
        badge.pack()
        badge_shell.place(relx=1.0, x=0, y=3, anchor="ne")

        body = ttk.Frame(surface, style="LaunchRoot.TFrame", padding=BODY_PADDING)
        body.grid(row=1, column=0, sticky="nsew")
        body.grid_columnconfigure(0, weight=1)
        panel_shell = tk.Frame(body, bg=colors["border"], padx=1, pady=1)
        panel_shell.grid(row=0, column=0, sticky="we")
        container = ttk.Frame(panel_shell, style="LaunchPanel.TFrame", padding=PANEL_PADDING)
        container.pack(fill="both", expand=True)
        container.grid_columnconfigure(1, weight=1)

        ttk.Label(container, text="Файли", style="LaunchSection.TLabel").grid(
            row=0, column=0, columnspan=3, sticky="w"
        )
        ttk.Label(
            container,
            text="Оберіть два Excel-файли та аркуші для порівняння.",
            style="LaunchMuted.TLabel",
            wraplength=520,
        ).grid(row=1, column=0, columnspan=3, pady=(3, 9), sticky="w")

        ttk.Label(container, text="Файл №1:", style="LaunchField.TLabel").grid(
            row=2, column=0, padx=(0, ROW_PAD_X), pady=ROW_PAD_Y, sticky="e"
        )
        self.file1_entry, self.sheet1_button = build_file_field(
            container,
            2,
            self.file1_display_var,
            self.choose_sheet1,
        )
        ttk.Button(container, text="ОБРАТИ", width=ACTION_BUTTON_WIDTH, style="Action.TButton", command=self.choose_file1).grid(
            row=2, column=2, pady=ROW_PAD_Y, sticky="w"
        )

        ttk.Label(container, text="Файл №2:", style="LaunchField.TLabel").grid(row=3, column=0, padx=(0, ROW_PAD_X), pady=ROW_PAD_Y, sticky="e")
        self.file2_entry, self.sheet2_button = build_file_field(
            container,
            3,
            self.file2_display_var,
            self.choose_sheet2,
        )
        ttk.Button(container, text="ОБРАТИ", width=ACTION_BUTTON_WIDTH, style="Action.TButton", command=self.choose_file2).grid(
            row=3, column=2, pady=ROW_PAD_Y, sticky="w"
        )

        ttk.Separator(container, orient="horizontal").grid(
            row=4, column=0, columnspan=3, sticky="we", pady=(11, 9)
        )

        self.compare_button = ttk.Button(
            container,
            text="ЗАПУСТИТИ",
            style="Launch.TButton",
            command=self.run_compare,
        )
        self.compare_button.grid(
            row=5, column=0, columnspan=3, pady=(0, 3), sticky="we"
        )

        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_rowconfigure(0, weight=1)

        self.file1_var.trace_add("write", lambda *_args: self._refresh_file_display(self.file1_var, self.file1_display_var, self.file1_entry))
        self.file2_var.trace_add("write", lambda *_args: self._refresh_file_display(self.file2_var, self.file2_display_var, self.file2_entry))
        self.file1_entry.bind("<Configure>", lambda _event: self._refresh_file_display(self.file1_var, self.file1_display_var, self.file1_entry))
        self.file2_entry.bind("<Configure>", lambda _event: self._refresh_file_display(self.file2_var, self.file2_display_var, self.file2_entry))
        self._refresh_file_display(self.file1_var, self.file1_display_var, self.file1_entry)
        self._refresh_file_display(self.file2_var, self.file2_display_var, self.file2_entry)
        self.root.update_idletasks()

    def _fit_path_for_entry(self, path: str, entry: ttk.Entry) -> str:
        path = path.strip()
        if not path:
            return ""

        normalized = Path(path)
        drive_prefix = ""
        if normalized.drive and len(normalized.drive) == 2 and normalized.drive[1] == ":":
            drive_prefix = f"{normalized.drive}/"

        file_name = normalized.name
        folder_name = normalized.parent.name
        if folder_name and file_name:
            display_path = f"{drive_prefix}.../{folder_name}/{file_name}"
        elif file_name:
            display_path = f"{drive_prefix}{file_name}" if drive_prefix else f".../{file_name}"
        elif drive_prefix:
            display_path = drive_prefix
        else:
            display_path = path.replace("\\", "/")

        entry.update_idletasks()
        font_name = entry.cget("font") or "TkDefaultFont"
        try:
            font = tkfont.nametofont(font_name)
        except tk.TclError:
            try:
                font = tkfont.nametofont("TkDefaultFont")
            except tk.TclError:
                return display_path
        available_width = max(120, entry.winfo_width() - 14)

        if font.measure(display_path) <= available_width:
            return display_path

        parts = display_path.split("/")
        if len(parts) >= 3:
            prefix = "/".join(parts[:-1]) + "/..."
            shortened_name = parts[-1]
            while len(shortened_name) > 1 and font.measure(prefix + shortened_name) > available_width:
                shortened_name = shortened_name[1:]
            return prefix + shortened_name

        ellipsis = "..."
        shortened_path = display_path
        while len(shortened_path) > 1 and font.measure(ellipsis + shortened_path) > available_width:
            shortened_path = shortened_path[1:]
        return ellipsis + shortened_path

    def _refresh_file_display(self, source_var: tk.StringVar, display_var: tk.StringVar, entry: ttk.Entry) -> None:
        if not widget_exists(entry):
            return

        path = source_var.get().strip()
        display_var.set(self._fit_path_for_entry(path, entry))

    def _set_sheet_button_enabled(self, button: tk.Button, enabled: bool) -> None:
        if not widget_exists(button):
            return

        try:
            button.configure(
                state="normal" if enabled else "disabled",
                cursor="hand2" if enabled else "arrow",
                fg=self.colors["text"] if enabled else self.colors["muted"],
                disabledforeground=self.colors["muted"],
            )
        except (tk.TclError, RuntimeError):
            pass

    def _set_compare_button_disabled(self, disabled: bool) -> None:
        if not widget_exists(self.compare_button):
            return
        try:
            self.compare_button.state(["disabled"] if disabled else ["!disabled"])
        except (tk.TclError, RuntimeError):
            pass

    def _schedule_queue_poll(self) -> None:
        if not widget_exists(self.root):
            return
        try:
            self.root.after(100, self._poll_compare_queue)
        except (tk.TclError, RuntimeError):
            pass

    def _load_sheets_for_path(self, path: str, file_label: str) -> Optional[List[str]]:
        cached = self.sheet_names_cache.get(path)
        if cached:
            return list(cached)

        try:
            sheet_names = get_sheet_names(path)
        except Exception as e:
            show_topmost_message("error", "Помилка", f"Не вдалося прочитати список аркушів у {file_label}:\n{e}", parent=self.root)
            return None

        if not sheet_names:
            show_topmost_message("warning", "Помилка", f"У {file_label} не знайдено жодного аркуша.", parent=self.root)
            return None

        self.sheet_names_cache[path] = list(sheet_names)
        return sheet_names

    def _invalidate_sheet_menu_cache(self, slot_prefix: str) -> None:
        keys_to_remove = [key for key in self.sheet_menu_cache if key[0] == slot_prefix]
        for key in keys_to_remove:
            menu = self.sheet_menu_cache.pop(key)
            try:
                menu.destroy()
            except Exception:
                pass

    def _build_sheet_menu(
        self,
        slot_prefix: str,
        target_sheet_var: tk.StringVar,
        sheet_names: List[str],
    ) -> tk.Menu:
        menu = tk.Menu(self.root, tearoff=0)

        def on_select(value: str) -> None:
            target_sheet_var.set(value)

        for name in sheet_names:
            menu.add_command(label=name, command=lambda value=name: on_select(value))

        return menu

    def _prime_sheet_menu_cache(
        self,
        slot_prefix: str,
        path: str,
        target_sheet_var: tk.StringVar,
        sheet_names: List[str],
    ) -> None:
        self._invalidate_sheet_menu_cache(slot_prefix)
        if not path or not sheet_names:
            return
        cache_key = (slot_prefix, path)
        self.sheet_menu_cache[cache_key] = self._build_sheet_menu(slot_prefix, target_sheet_var, sheet_names)

    def choose_file1(self):
        prepare_dialog_parent(self.root)
        path = filedialog.askopenfilename(
            title="Оберіть Файл №1",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")],
            parent=self.root,
        )
        if not path:
            return

        self.file1_var.set(path)
        self.sheet1_var.set("")
        self.file1_sheet_names = []
        self._invalidate_sheet_menu_cache("file1")
        self._set_sheet_button_enabled(self.sheet1_button, False)
        self.root.update_idletasks()
        sheets = self._load_sheets_for_path(path, "Файлі №1")
        if sheets:
            self.file1_sheet_names = sheets
            self.sheet1_var.set(sheets[0])
            self._prime_sheet_menu_cache("file1", path, self.sheet1_var, sheets)
            self._set_sheet_button_enabled(self.sheet1_button, True)
        else:
            self.file1_sheet_names = []
            self._set_sheet_button_enabled(self.sheet1_button, False)
        prepare_dialog_parent(self.root)

    def choose_file2(self):
        prepare_dialog_parent(self.root)
        path = filedialog.askopenfilename(
            title="Оберіть Файл №2",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")],
            parent=self.root,
        )
        if not path:
            return

        self.file2_var.set(path)
        self.sheet2_var.set("")
        self.file2_sheet_names = []
        self._invalidate_sheet_menu_cache("file2")
        self._set_sheet_button_enabled(self.sheet2_button, False)
        self.root.update_idletasks()
        sheets = self._load_sheets_for_path(path, "Файлі №2")
        if sheets:
            self.file2_sheet_names = sheets
            self.sheet2_var.set(sheets[0])
            self._prime_sheet_menu_cache("file2", path, self.sheet2_var, sheets)
            self._set_sheet_button_enabled(self.sheet2_button, True)
        else:
            self.file2_sheet_names = []
            self._set_sheet_button_enabled(self.sheet2_button, False)
        prepare_dialog_parent(self.root)

    def choose_sheet1(self):
        self._choose_sheet(
            "file1",
            self.file1_var,
            self.sheet1_var,
            self.sheet1_button,
            self.file1_sheet_names,
            "Файлу №1",
        )

    def choose_sheet2(self):
        self._choose_sheet(
            "file2",
            self.file2_var,
            self.sheet2_var,
            self.sheet2_button,
            self.file2_sheet_names,
            "Файлу №2",
        )

    def _choose_sheet(
        self,
        slot_prefix: str,
        file_var: tk.StringVar,
        target_sheet_var: tk.StringVar,
        anchor_button: ttk.Button,
        cached_sheet_names: List[str],
        file_label: str,
    ):
        path = file_var.get().strip()
        if not path or not os.path.exists(path):
            show_topmost_message("warning", "Перевірка", f"Спершу оберіть коректний файл для {file_label}.", parent=self.root)
            return

        sheet_names = cached_sheet_names
        if not sheet_names:
            sheet_names = self._load_sheets_for_path(path, file_label)
            if not sheet_names:
                return
            if slot_prefix == "file1":
                self.file1_sheet_names = sheet_names
            else:
                self.file2_sheet_names = sheet_names
            self._prime_sheet_menu_cache(slot_prefix, path, target_sheet_var, sheet_names)

        cache_key = (slot_prefix, path)
        menu = self.sheet_menu_cache.get(cache_key)
        if menu is None:
            self._prime_sheet_menu_cache(slot_prefix, path, target_sheet_var, sheet_names)
            menu = self.sheet_menu_cache.get(cache_key)
        if menu is None:
            return

        selected_name = target_sheet_var.get().strip()
        for index, name in enumerate(sheet_names):
            label = f"✓ {name}" if name == selected_name else name
            menu.entryconfigure(index, label=label)

        self.root.update_idletasks()
        x = anchor_button.winfo_rootx()
        y = anchor_button.winfo_rooty() + anchor_button.winfo_height()
        try:
            menu.tk_popup(x, y)
        finally:
            menu.grab_release()

    def run_compare(self):
        if self.compare_in_progress:
            return

        file1 = self.file1_var.get().strip()
        file2 = self.file2_var.get().strip()
        sheet1 = self.sheet1_var.get().strip()
        sheet2 = self.sheet2_var.get().strip()

        if not file1 or not os.path.exists(file1):
            show_topmost_message("warning", "Перевірка", "Оберіть коректний Файл №1.", parent=self.root)
            return

        if not file2 or not os.path.exists(file2):
            show_topmost_message("warning", "Перевірка", "Оберіть коректний Файл №2.", parent=self.root)
            return

        if not sheet1:
            show_topmost_message("warning", "Перевірка", "Оберіть аркуш для Файлу №1.", parent=self.root)
            return

        if not sheet2:
            show_topmost_message("warning", "Перевірка", "Оберіть аркуш для Файлу №2.", parent=self.root)
            return

        while True:
            try:
                self.compare_queue.get_nowait()
            except Empty:
                break

        self._initialize_process_log(file1, file2, sheet1, sheet2)
        self._begin_run_timing()
        self._append_process_log("Валідацію введених координат пройдено")

        self.compare_in_progress = True
        self._set_compare_button_disabled(True)
        try:
            self.root.grab_release()
        except Exception:
            pass
        self.root.withdraw()
        self.progress_window = ProgressWindow(self.root)
        self.progress_window.update(
            header="Етап 1/4: Підготовка",
            detail="Відкриваємо книги Excel та перевіряємо аркуші.",
            file_name=f"{os.path.basename(file1)} <> {os.path.basename(file2)}",
            current=0,
            total=4,
        )

        worker = threading.Thread(
            target=self._compare_worker,
            args=(file1, file2, sheet1, sheet2),
            daemon=True,
        )
        worker.start()
        self._schedule_queue_poll()

    def _compare_worker(self, file1: str, file2: str, sheet1: str, sheet2: str):
        try:
            self._begin_stage("stage1", "Етап 1/4: відкриття книг Excel")
            loaded1 = load_workbook_generic(file1)
            loaded2 = load_workbook_generic(file2)
            self._append_process_log(f"Книги відкрито: {os.path.basename(file1)} | {os.path.basename(file2)}")
            self._finish_stage("stage1", "Етап 1/4: відкриття книг Excel")
            self.compare_queue.put(("progress", "Етап 2/4: Аналіз", "Аналізуємо вибрані аркуші та шукаємо потрібні колонки.", "", 1, 4))

            self._begin_stage("stage2", "Етап 2/4: пошук колонок серійників і цін")
            ctx1 = resolve_column_context(
                loaded=loaded1,
                sheet_name=sheet1,
                logical_name="serial_number",
                column_aliases=COLUMN_ALIASES,
            )

            ctx2 = resolve_column_context(
                loaded=loaded2,
                sheet_name=sheet2,
                logical_name="serial_number",
                column_aliases=COLUMN_ALIASES,
            )

            price_ctx1 = resolve_price_source_context(loaded1, sheet1)
            price_ctx2 = resolve_price_source_context(loaded2, sheet2)
            self._append_process_log(
                f"Файл №1: serial_col={ctx1.serial_col_index}, data_start_row={ctx1.header_data_start_row}, "
                f"price_source={price_ctx1.source_kind}, value_col={price_ctx1.value_col_index}, "
                f"quantity_col={price_ctx1.quantity_col_index}, price_header='{price_ctx1.matched_alias}'"
            )
            self._append_process_log(
                f"Файл №2: serial_col={ctx2.serial_col_index}, data_start_row={ctx2.header_data_start_row}, "
                f"price_source={price_ctx2.source_kind}, value_col={price_ctx2.value_col_index}, "
                f"quantity_col={price_ctx2.quantity_col_index}, price_header='{price_ctx2.matched_alias}'"
            )
            self._finish_stage("stage2", "Етап 2/4: пошук колонок серійників і цін")

            self.compare_queue.put(("progress", "Етап 3/4: Серійники", "Зчитуємо серійники та формуємо первинний звіт.", "", 2, 4))
            self._begin_stage("stage3", "Етап 3/4: зчитування серійників")
            serial_occurrences1 = collect_serial_occurrences(loaded1, ctx1)
            serial_occurrences2 = collect_serial_occurrences(loaded2, ctx2)
            serials1 = set(serial_occurrences1.keys())
            serials2 = set(serial_occurrences2.keys())
            duplicate_serials = build_duplicate_serial_evidences(serial_occurrences1, serial_occurrences2)

            only_in_file1 = sorted(serials1 - serials2)
            only_in_file2 = sorted(serials2 - serials1)
            in_both = sorted(serials1 & serials2)
            self._append_process_log(
                f"Серійники: file1={len(serials1)}, file2={len(serials2)}, "
                f"only1={len(only_in_file1)}, only2={len(only_in_file2)}, "
                f"both={len(in_both)}, duplicates={len(duplicate_serials)}"
            )

            out_path0, out_path1, out_path2, out_path3 = save_reports_near_file2(
                file1_path=file1,
                file2_path=file2,
                sheet1_name=sheet1,
                sheet2_name=sheet2,
                duplicate_serials=duplicate_serials,
                only_in_file1=only_in_file1,
                only_in_file2=only_in_file2,
                in_both=in_both,
            )
            self._append_process_log(
                f"Збережено логи серійників: {out_path0} | {out_path1} | {out_path2} | {out_path3}"
            )
            self._finish_stage("stage3", "Етап 3/4: зчитування серійників")

            self.compare_queue.put(("progress", "Етап 4/4: Ціни", "Звіряємо ціни для спільних серійників.", "", 0, max(len(in_both), 1)))
            self._begin_stage("stage4", "Етап 4/4: звірка цін і позначення колізій")

            def push_price_progress(idx: int, total: int, serial: str) -> None:
                if total <= 50 or idx in (1, total) or idx % 25 == 0:
                    self.compare_queue.put(
                        (
                            "progress",
                            "Етап 4/4: Ціни",
                            f"Звіряємо ціни: {idx} / {total}",
                            serial,
                            idx,
                            total,
                        )
                    )

            matched_prices, mismatched_prices = compare_prices_for_common_serials(
                loaded1=loaded1,
                loaded2=loaded2,
                sheet1=sheet1,
                sheet2=sheet2,
                serial_ctx1=ctx1,
                serial_ctx2=ctx2,
                price_ctx1=price_ctx1,
                price_ctx2=price_ctx2,
                serials=in_both,
                progress_callback=push_price_progress,
            )

            out_path4, out_path5 = save_price_reports_near_file2(
                file1_path=file1,
                file2_path=file2,
                sheet1_name=sheet1,
                sheet2_name=sheet2,
                matched=matched_prices,
                mismatched=mismatched_prices,
            )
            self._append_process_log(
                f"Збережено логи цін: {out_path4} | {out_path5}"
            )
            self._append_process_log(
                f"Результат звірки цін: matched={len(matched_prices)}, mismatched={len(mismatched_prices)}"
            )
            self.compare_queue.put(
                (
                    "progress",
                    "Етап 4/4: Позначення",
                    "Фарбуємо рядки з колізіями та додаємо примітки у вихідні книги.",
                    "",
                    max(len(in_both), 1),
                    max(len(in_both), 1),
                )
            )
            annotation_messages = annotate_mismatched_price_rows(
                file1_path=file1,
                file2_path=file2,
                sheet1_name=sheet1,
                sheet2_name=sheet2,
                serial_ctx1=ctx1,
                serial_ctx2=ctx2,
                evidences=mismatched_prices,
            )
            for message in annotation_messages:
                self._append_process_log(message)
            self._finish_stage("stage4", "Етап 4/4: звірка цін і позначення колізій")
            self.compare_queue.put(
                (
                    "success",
                    out_path0,
                    out_path1,
                    out_path2,
                    out_path3,
                    out_path4,
                    out_path5,
                    len(duplicate_serials),
                    len(only_in_file1),
                    len(only_in_file2),
                    len(in_both),
                    len(matched_prices),
                    len(mismatched_prices),
                    annotation_messages,
                )
            )
        except Exception as e:
            for stage_key, stage_message in (
                ("stage1", "Етап 1/4: відкриття книг Excel"),
                ("stage2", "Етап 2/4: пошук колонок серійників і цін"),
                ("stage3", "Етап 3/4: зчитування серійників"),
                ("stage4", "Етап 4/4: звірка цін і позначення колізій"),
            ):
                self._finish_stage(stage_key, stage_message, outcome="перервано після")
            self._append_process_log(f"ERROR: {e}")
            self.compare_queue.put(("error", getattr(e, "title", "Помилка"), str(e)))

    def _poll_compare_queue(self):
        if not widget_exists(self.root):
            self.compare_in_progress = False
            if self.progress_window is not None:
                self.progress_window.close()
                self.progress_window = None
            return

        latest_progress = None
        terminal_result = None

        while True:
            try:
                result = self.compare_queue.get_nowait()
            except Empty:
                break

            if result[0] == "progress":
                latest_progress = result
                continue

            terminal_result = result
            break

        if terminal_result is None and latest_progress is None:
            if self.compare_in_progress:
                self._schedule_queue_poll()
            return

        if terminal_result is None and latest_progress is not None:
            _, header, detail, file_name, current, total = latest_progress
            if self.progress_window is not None:
                self.progress_window.update(
                    header=header,
                    detail=detail,
                    file_name=file_name,
                    current=current,
                    total=total,
                )
            self._schedule_queue_poll()
            return

        result = terminal_result
        self.compare_in_progress = False
        self._set_compare_button_disabled(False)

        error_parent = self.root
        if self.progress_window is not None and widget_exists(self.progress_window.dialog):
            error_parent = self.progress_window.dialog

        if result[0] == "success":
            (
                _,
                out_path0,
                out_path1,
                out_path2,
                out_path3,
                out_path4,
                out_path5,
                duplicate_count,
                only1_count,
                only2_count,
                both_count,
                matched_count,
                mismatched_count,
                annotation_messages,
            ) = result
            annotation_text = ""
            if annotation_messages:
                annotation_text = "\n\nПозначки у книгах:\n" + "\n".join(
                    f"- {message}" for message in annotation_messages
                )
            success_text = (
                f"Порівняння завершено.\n\n"
                f"Результати збережено у файлах:\n"
                f"{out_path0}\n"
                f"{out_path1}\n"
                f"{out_path2}\n"
                f"{out_path3}\n"
                f"{out_path4}\n"
                f"{out_path5}\n\n"
                f"Знайдено:\n"
                f"- серійники з дублями: {duplicate_count}\n"
                f"- тільки у Файлі №1: {only1_count}\n"
                f"- тільки у Файлі №2: {only2_count}\n"
                f"- у обох файлах: {both_count}\n"
                f"- ціни збігаються: {matched_count}\n"
                f"- ціни НЕ збігаються: {mismatched_count}"
                f"{annotation_text}"
            )
            if self.progress_window is not None:
                self.progress_window.close()
                self.progress_window = None
            show_topmost_message(
                "info",
                "Готово",
                success_text,
                parent=self.root,
                reveal_parent=False,
            )
            self._append_process_log("SUCCESS: порівняння завершено")
            self._append_process_log(success_text.replace("\n", " | "))
            self._finish_run_timing()
            destroy_widget(self.root)
            return

        if len(result) == 3:
            _, error_title, error_text = result
        else:
            _, error_text = result
            error_title = "Помилка"
        self._append_process_log(f"FAIL: {error_text}")
        self._finish_run_timing()
        show_topmost_message("error", error_title, error_text, parent=error_parent)
        if self.progress_window is not None:
            self.progress_window.close()
            self.progress_window = None
        destroy_widget(self.root)


# ------------------------------------------------------------
# ENTRY POINT
# ------------------------------------------------------------

def main():
    root = tk.Tk()
    root.withdraw()
    install_frozen_executable_icon(root)
    app = App(root)
    root.update_idletasks()
    root.deiconify()
    root.lift()
    try:
        root.grab_set()
        root.focus_force()
    except tk.TclError:
        pass
    root.mainloop()


if __name__ == "__main__":
    main()
